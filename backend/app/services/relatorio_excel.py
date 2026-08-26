# -*- coding: utf-8 -*-
from __future__ import annotations

import json
import re
from collections import Counter
from datetime import datetime
from email.utils import parsedate_to_datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from persistencia.db import conectar, init_db

BLUE = "2E3192"
GREEN = "DFF3E6"
YELLOW = "FFF0C2"
RED = "FDE2E1"
GRAY = "EEF2F7"
TEXT = "1F2937"


def _json_list(valor: str | None) -> list[str]:
    if not valor:
        return []
    try:
        parsed = json.loads(valor)
        return parsed if isinstance(parsed, list) else []
    except json.JSONDecodeError:
        return []


def _fmt_data(valor: str | None) -> str:
    if not valor:
        return ""
    try:
        return datetime.fromisoformat(valor).strftime("%d/%m/%Y %H:%M")
    except ValueError:
        try:
            return parsedate_to_datetime(valor).strftime("%d/%m/%Y %H:%M")
        except (TypeError, ValueError):
            return valor


def _texto_usuario(valor: str) -> str:
    texto = str(valor or "").strip()
    if texto in {"None", "none", "NULL", "null"}:
        return "em branco"
    if len(texto) >= 2 and texto[0] == texto[-1] and texto[0] in {"'", '"'}:
        texto = texto[1:-1]

    match_dt = re.fullmatch(
        r"datetime\.datetime\((\d{4}),\s*(\d{1,2}),\s*(\d{1,2})(?:,\s*(\d{1,2}),\s*(\d{1,2}))?.*\)",
        texto,
    )
    if match_dt:
        ano, mes, dia, hora, minuto = match_dt.groups()
        data = datetime(
            int(ano),
            int(mes),
            int(dia),
            int(hora or 0),
            int(minuto or 0),
        )
        return data.strftime("%d/%m/%Y %H:%M") if hora else data.strftime("%d/%m/%Y")

    return texto or "em branco"


def _parse_evidencia(texto: str, tipo: str) -> dict[str, str]:
    texto = str(texto or "").strip()
    padroes = [
        r'^(.*?): antes "([\s\S]*)"; depois "([\s\S]*)"$',
        r"^(.*?): antes ([\s\S]*); depois ([\s\S]*)$",
        r"^(.*?): (linha anterior.*?); antes \(([\s\S]*)\); depois \(([\s\S]*)\)$",
    ]
    for padrao in padroes:
        match = re.match(padrao, texto)
        if not match:
            continue
        if len(match.groups()) == 4:
            return {
                "local": f"{match.group(1)} - {match.group(2)}",
                "antes": _texto_usuario(match.group(3)),
                "depois": _texto_usuario(match.group(4)),
            }
        return {
            "local": match.group(1),
            "antes": _texto_usuario(match.group(2)),
            "depois": _texto_usuario(match.group(3)),
        }

    incluido = re.match(r'^(.*?): incluído "([\s\S]*)"$', texto)
    if incluido:
        return {"local": incluido.group(1), "antes": "", "depois": _texto_usuario(incluido.group(2))}
    removido = re.match(r'^(.*?): removido "([\s\S]*)"$', texto)
    if removido:
        return {"local": removido.group(1), "antes": _texto_usuario(removido.group(2)), "depois": ""}
    interno = re.match(r'^Arquivo interno incluído: ([^;]+); evidência: "([\s\S]*)"$', texto)
    if interno:
        return {"local": f"Arquivo interno {interno.group(1)}", "antes": "", "depois": _texto_usuario(interno.group(2))}

    if tipo == "Saiu":
        return {"local": "Evidência", "antes": _texto_usuario(texto), "depois": ""}
    return {"local": "Evidência", "antes": "", "depois": _texto_usuario(texto)}


def _grupo_local(local: str) -> str:
    pagina = re.match(r"^(Página \d+)", local)
    if pagina:
        return pagina.group(1)
    arquivo = re.match(r"^(Arquivo interno [^-:]+)", local)
    if arquivo:
        return arquivo.group(1).strip()
    if re.match(r"^linha (atual|anterior) \d+", local, flags=re.I):
        return "Linhas do arquivo"
    return local


def _numero_linha(local: str, tipo: str) -> int | None:
    padrao = r"linha anterior (\d+)" if tipo == "Saiu" else r"linha atual (\d+)"
    match = re.search(padrao, local, flags=re.I)
    return int(match.group(1)) if match else None


def _numero_linha_por_padrao(local: str, padrao: str) -> int | None:
    match = re.search(padrao, local, flags=re.I)
    return int(match.group(1)) if match else None


def _titulo_local_grupo(grupo: str, numeros: list[int]) -> str:
    if not numeros:
        return grupo
    numeros = sorted(numeros)
    if len(numeros) == 1:
        return f"{grupo} - linha {numeros[0]}"
    return f"{grupo} - linhas {numeros[0]} a {numeros[-1]}"


def _titulo_local_mudou(grupo: str, dados: list[dict[str, str]]) -> str:
    anteriores = [
        numero
        for item in dados
        if (numero := _numero_linha_por_padrao(item["local"], r"linha anterior (\d+)")) is not None
    ]
    atuais = [
        numero
        for item in dados
        if (numero := _numero_linha_por_padrao(item["local"], r"linha atual (\d+)")) is not None
    ]
    if not anteriores and not atuais:
        return grupo
    if len(anteriores) <= 1 and len(atuais) <= 1:
        return dados[0]["local"]
    trecho_anterior = (
        f"linhas anteriores {min(anteriores)} a {max(anteriores)}"
        if len(anteriores) > 1
        else f"linha anterior {anteriores[0]}"
    )
    trecho_atual = (
        f"linhas atuais {min(atuais)} a {max(atuais)}"
        if len(atuais) > 1
        else f"linha atual {atuais[0]}"
    )
    return f"{grupo} - {trecho_anterior} -> {trecho_atual}"


def _itens_reais(itens: list[str]) -> list[str]:
    return [item for item in itens if not str(item).strip().startswith("...")]


def _linhas_evidencia_agrupadas(itens: list[str], tipo: str) -> list[dict[str, str]]:
    grupos: list[tuple[str, list[dict[str, str]]]] = []
    for texto in itens:
        parsed = _parse_evidencia(texto, tipo)
        grupo = _grupo_local(parsed["local"])
        if grupos and grupos[-1][0] == grupo:
            grupos[-1][1].append(parsed)
        else:
            grupos.append((grupo, [parsed]))

    agrupados: list[dict[str, str]] = []
    for grupo, dados in grupos:
        if tipo == "Mudou":
            agrupados.append(
                {
                    "local": _titulo_local_mudou(grupo, dados),
                    "antes": "\n".join(
                        item.get("antes", "").strip()
                        for item in dados
                        if item.get("antes", "").strip()
                    ),
                    "depois": "\n".join(
                        item.get("depois", "").strip()
                        for item in dados
                        if item.get("depois", "").strip()
                    ),
                }
            )
            continue

        numeros = [
            numero
            for item in dados
            if (numero := _numero_linha(item["local"], tipo)) is not None
        ]
        campo_texto = "antes" if tipo == "Saiu" else "depois"
        textos = [item.get(campo_texto, "").strip() for item in dados if item.get(campo_texto, "").strip()]
        agrupados.append(
            {
                "local": _titulo_local_grupo(grupo, numeros),
                "antes": "\n".join(textos) if tipo == "Saiu" else "",
                "depois": "\n".join(textos) if tipo != "Saiu" else "",
            }
        )
    return agrupados


def _linhas_pdf(caminho: str | None, cache: dict[str, list[list[tuple[int, str]]]]) -> list[list[tuple[int, str]]]:
    if not caminho:
        return []
    if caminho in cache:
        return cache[caminho]

    path = Path(caminho)
    if not path.exists():
        cache[caminho] = []
        return []

    try:
        from pypdf import PdfReader  # type: ignore

        reader = PdfReader(str(path))
        paginas: list[list[tuple[int, str]]] = []
        for page in reader.pages:
            linhas = []
            for numero, linha in enumerate((page.extract_text() or "").splitlines(), start=1):
                normalizada = re.sub(r"\s+", " ", linha).strip()
                if normalizada:
                    linhas.append((numero, normalizada))
            paginas.append(linhas)
        cache[caminho] = paginas
        return paginas
    except Exception:
        cache[caminho] = []
        return []


def _contexto_pdf(
    caminho: str | None,
    local: str,
    cache: dict[str, list[list[tuple[int, str]]]],
    *,
    raio: int = 5,
) -> tuple[str, str] | None:
    match_pagina = re.search(r"Página (\d+)", local)
    match_linha = re.search(r"linha(?: atual| anterior)? (\d+)", local, flags=re.I)
    if not match_pagina or not match_linha:
        return None

    pagina_idx = int(match_pagina.group(1)) - 1
    linha_ref = int(match_linha.group(1))
    paginas = _linhas_pdf(caminho, cache)
    if pagina_idx < 0 or pagina_idx >= len(paginas):
        return None

    linhas = [
        (numero, texto)
        for numero, texto in paginas[pagina_idx]
        if linha_ref - raio <= numero <= linha_ref + raio
    ]
    if not linhas:
        return None

    inicio = linhas[0][0]
    fim = linhas[-1][0]
    titulo = f"Página {pagina_idx + 1} - linha {linha_ref} (contexto linhas {inicio} a {fim})"
    texto = "\n".join(texto for _, texto in linhas)
    return titulo, texto


def _enriquecer_contexto_pdf(
    parsed: dict[str, str],
    tipo: str,
    alt: dict[str, Any],
    cache_pdf: dict[str, list[list[tuple[int, str]]]],
) -> dict[str, str]:
    if str(alt.get("tipo_arquivo") or "").lower() != "pdf":
        return parsed
    if not re.search(r"Página \d+ - linha \d+$", parsed["local"]):
        return parsed

    enriched = dict(parsed)
    if tipo == "Entrou":
        contexto = _contexto_pdf(alt.get("versao_atual"), parsed["local"], cache_pdf)
        if contexto:
            enriched["local"], enriched["depois"] = contexto
    elif tipo == "Saiu":
        contexto = _contexto_pdf(alt.get("versao_anterior"), parsed["local"], cache_pdf)
        if contexto:
            enriched["local"], enriched["antes"] = contexto
    elif tipo == "Mudou":
        contexto_ant = _contexto_pdf(alt.get("versao_anterior"), parsed["local"], cache_pdf)
        contexto_novo = _contexto_pdf(alt.get("versao_atual"), parsed["local"], cache_pdf)
        if contexto_novo:
            enriched["local"], enriched["depois"] = contexto_novo
        if contexto_ant:
            enriched["antes"] = contexto_ant[1]
    return enriched


def _buscar_alteracoes(escopo: str) -> tuple[list[dict[str, Any]], str]:
    init_db()
    where = ""
    params: list[Any] = []
    titulo = "Histórico de alterações de leiautes Bacen"

    with conectar() as conn:
        if escopo == "ultima":
            row = conn.execute(
                """
                SELECT execucao_id
                FROM alteracoes_detectadas
                ORDER BY execucao_id DESC, id DESC
                LIMIT 1
                """
            ).fetchone()
            if row:
                where = "WHERE a.execucao_id = ?"
                params.append(row["execucao_id"])
                titulo = "Relatório de alterações de leiautes Bacen"

        rows = conn.execute(
            f"""
            SELECT
                a.id,
                a.execucao_id,
                a.status,
                a.criado_em,
                a.resumo_executivo,
                a.impacto_sugerido,
                a.itens_incluidos,
                a.itens_removidos,
                a.itens_alterados,
                COALESCE(l.codigo, '') AS leiaute_codigo,
                COALESCE(l.nome, '') AS leiaute_nome,
                ar.nome_arquivo,
                ar.tipo_arquivo,
                ar.last_modified,
                ar.url,
                ar.final_url,
                e.iniciado_em,
                e.finalizado_em,
                va.caminho_arquivo AS versao_anterior,
                vn.caminho_arquivo AS versao_atual,
                va.tamanho_bytes AS tamanho_anterior,
                vn.tamanho_bytes AS tamanho_atual
            FROM alteracoes_detectadas a
            JOIN arquivos_monitorados ar ON ar.id = a.arquivo_id
            LEFT JOIN leiautes_monitorados l ON l.id = ar.leiaute_id
            LEFT JOIN execucoes e ON e.id = a.execucao_id
            LEFT JOIN versoes_arquivos va ON va.id = a.versao_anterior_id
            LEFT JOIN versoes_arquivos vn ON vn.id = a.versao_atual_id
            {where}
            ORDER BY a.execucao_id DESC, a.id DESC
            """,
            params,
        ).fetchall()

    alteracoes = []
    for row in rows:
        item = dict(row)
        item["itens_incluidos"] = _json_list(item.get("itens_incluidos"))
        item["itens_removidos"] = _json_list(item.get("itens_removidos"))
        item["itens_alterados"] = _json_list(item.get("itens_alterados"))
        alteracoes.append(item)
    return alteracoes, titulo


def _append(ws, values: list[Any], fill: str | None = None, bold: bool = False) -> None:
    ws.append(values)
    for cell in ws[ws.max_row]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.font = Font(name="Arial", size=10, bold=bold, color=TEXT)
        cell.border = Border(bottom=Side(style="thin", color="E5E7EB"))
        if fill:
            cell.fill = PatternFill("solid", fgColor=fill)


def _titulo(ws, texto: str, subtitulo: str) -> None:
    ws["A1"] = texto
    ws["A1"].font = Font(name="Arial", size=18, bold=True, color=BLUE)
    ws["A2"] = subtitulo
    ws["A2"].font = Font(name="Arial", size=10, color="64748B")
    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")


def _aplicar_cor_linha(ws, fill: str) -> None:
    for cell in ws[ws.max_row]:
        cell.fill = PatternFill("solid", fgColor=fill)


def _autofiltro(ws, primeira_linha: int, ultima_coluna: int) -> None:
    if ws.max_row <= primeira_linha:
        return
    ref = f"A{primeira_linha}:{get_column_letter(ultima_coluna)}{ws.max_row}"
    ws.freeze_panes = f"A{primeira_linha + 1}"
    ws.auto_filter.ref = ref


def _ajustar(ws, larguras: dict[str, int]) -> None:
    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 24
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 38
    ws.sheet_view.showGridLines = False


def _altura_dados(ws, altura: int) -> None:
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = altura


def _metric_cell(cell: Cell, fill: str) -> None:
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name="Arial", size=14, bold=True, color=TEXT)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin", color="D8DEE8"),
        right=Side(style="thin", color="D8DEE8"),
        top=Side(style="thin", color="D8DEE8"),
        bottom=Side(style="thin", color="D8DEE8"),
    )


def _eh_evidencia_tecnica_texto(texto: str) -> bool:
    t = (texto or "").lower()
    if "novo arquivo observado" in t or "arquivo novo" in t:
        return False
    chaves = (
        "etag",
        "last_modified",
        "content_length",
        "final_url",
        "partial_fp",
        "metadados",
        "versão anterior não arquivada",
        "versao anterior nao arquivada",
        "alteracao detectada por metadados",
        "alteração detectada por metadados",
    )
    return any(k in t for k in chaves)


def _alteracao_so_aviso(alt: dict[str, Any]) -> bool:
    """Republicação sem mudança de conteúdo — mesma ideia do e-mail."""
    resumo = str(alt.get("resumo_executivo") or "")
    if "novo arquivo" in resumo.lower() or "versão pareada" in resumo.lower():
        return False
    inc = _itens_reais(alt.get("itens_incluidos") or [])
    rem = _itens_reais(alt.get("itens_removidos") or [])
    alt_itens = _itens_reais(alt.get("itens_alterados") or [])
    conteudo_alt = [x for x in alt_itens if not _eh_evidencia_tecnica_texto(str(x))]
    if inc or rem or conteudo_alt:
        return False
    todos = list(inc) + list(rem) + list(alt_itens) + [resumo]
    return bool(todos) and all(_eh_evidencia_tecnica_texto(str(x)) for x in todos if str(x).strip())


def _codigo_situacao(resumo: str) -> str:
    if resumo.startswith("[Versão pareada") or resumo.startswith(
        "[Comparado com a versão anterior"
    ):
        return "versao_pareada"
    if resumo.startswith("[Sem anterior]") or resumo.startswith("[Arquivo novo]"):
        return "sem_anterior"
    if resumo.startswith("[Mesmo arquivo]"):
        return "mesmo_arquivo"
    return "desconhecido"


def _o_que_mudou_de_parsed(parsed: dict[str, str], tipo: str) -> str:
    mudanca = str(parsed.get("mudanca") or "").strip()
    if mudanca:
        return mudanca
    local = str(parsed.get("local") or "").lower()
    if "aba" in local and tipo == "Entrou":
        return "acrescentou aba"
    if "aba" in local and tipo == "Saiu":
        return "removeu aba"
    if tipo == "Entrou":
        return "conteúdo novo"
    if tipo == "Saiu":
        return "conteúdo removido"
    return "texto alterado"


def gerar_relatorio_alteracoes_xlsx(escopo: str = "historico") -> tuple[bytes, str]:
    """Exportar do Histórico — mesma planilha do anexo do e-mail (linguagem simples)."""
    from persistencia.planilha_gestor import (
        ArquivoResumoPlanilha,
        DadosPlanilhaGestor,
        LinhaMudancaPlanilha,
        gerar_bytes_planilha_gestor,
        rotulo_situacao,
    )

    escopo = "ultima" if escopo == "ultima" else "historico"
    alteracoes, _titulo = _buscar_alteracoes(escopo)
    dados = DadosPlanilhaGestor()
    cache_pdf: dict[str, list[list[tuple[int, str]]]] = {}

    for alt in alteracoes:
        data = _fmt_data(alt.get("last_modified")) or _fmt_data(alt.get("criado_em"))
        leiaute = str(alt.get("leiaute_codigo") or "")
        arquivo = str(alt.get("nome_arquivo") or "")
        link = str(alt.get("final_url") or alt.get("url") or "")
        o_que_fazer = str(alt.get("impacto_sugerido") or "Revisar o arquivo no site do Bacen.")
        resumo = str(alt.get("resumo_executivo") or "")

        if _alteracao_so_aviso(alt):
            dados.arquivos_aviso.append(
                ArquivoResumoPlanilha(
                    data=data,
                    leiaute=leiaute,
                    arquivo=arquivo,
                    situacao=rotulo_situacao("aviso"),
                    precisa_agir=False,
                    qtd_mudancas=0,
                    link=link,
                    observacao=(
                        "O Bacen republicou o arquivo no site, sem mudança de "
                        "texto, célula ou tabela."
                    ),
                )
            )
            continue

        linhas_arquivo: list[LinhaMudancaPlanilha] = []
        for tipo, itens in [
            ("Entrou", alt["itens_incluidos"]),
            ("Saiu", alt["itens_removidos"]),
            ("Mudou", alt["itens_alterados"]),
        ]:
            itens_validos = [
                x
                for x in _itens_reais(itens)
                if not _eh_evidencia_tecnica_texto(str(x))
            ]
            for parsed in _linhas_evidencia_agrupadas(itens_validos, tipo):
                parsed = _enriquecer_contexto_pdf(parsed, tipo, alt, cache_pdf)
                linhas_arquivo.append(
                    LinhaMudancaPlanilha(
                        data=data,
                        leiaute=leiaute,
                        arquivo=arquivo,
                        onde=str(parsed.get("local") or "—"),
                        o_que_mudou=_o_que_mudou_de_parsed(parsed, tipo),
                        antes=str(parsed.get("antes") or "—") or "—",
                        depois=str(parsed.get("depois") or "—") or "—",
                        o_que_fazer=o_que_fazer,
                    )
                )

        if not linhas_arquivo:
            linhas_arquivo.append(
                LinhaMudancaPlanilha(
                    data=data,
                    leiaute=leiaute,
                    arquivo=arquivo,
                    onde="—",
                    o_que_mudou="Arquivo novo ou atualizado sem detalhe de células",
                    antes="—",
                    depois="—",
                    o_que_fazer=o_que_fazer,
                )
            )

        dados.arquivos_agir.append(
            ArquivoResumoPlanilha(
                data=data,
                leiaute=leiaute,
                arquivo=arquivo,
                situacao=rotulo_situacao(_codigo_situacao(resumo)),
                precisa_agir=True,
                qtd_mudancas=len(linhas_arquivo),
                link=link,
            )
        )
        dados.linhas_mudanca.extend(linhas_arquivo)

    sufixo = "envio" if escopo == "ultima" else "historico"
    nome = f"mudancas_leiautes_{sufixo}_{datetime.now().strftime('%d%m%Y_%H%M')}.xlsx"
    return gerar_bytes_planilha_gestor(dados, nome_arquivo=nome)
