# -*- coding: utf-8 -*-
from __future__ import annotations

import difflib
import hashlib
import re
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

RAIZ = Path(__file__).resolve().parent.parent.parent.parent


def _resolver(caminho: str | None) -> Path | None:
    if not caminho:
        return None
    path = Path(caminho)
    if not path.is_absolute():
        path = RAIZ / path
    return path


def _limitar(itens: list[str], limite: int = 80) -> list[str]:
    if len(itens) <= limite:
        return itens
    return [*itens[:limite], f"... lista completa excede {limite} item(ns)"]


def _ler_texto(path: Path) -> str:
    data = path.read_bytes()
    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="ignore")


def _normalizar_linhas(texto: str) -> list[str]:
    linhas = []
    for linha in texto.splitlines():
        normalizada = re.sub(r"\s+", " ", linha).strip()
        if normalizada:
            linhas.append(normalizada)
    return linhas


def _linhas_com_numero(texto: str) -> list[tuple[int, str]]:
    linhas = []
    for numero, linha in enumerate(texto.splitlines(), start=1):
        normalizada = re.sub(r"\s+", " ", linha).strip()
        if normalizada:
            linhas.append((numero, normalizada))
    return linhas


def _formatar_trecho(valor: str, limite: int = 5000) -> str:
    valor = valor.strip()
    if len(valor) <= limite:
        return valor
    return f"{valor[:limite]}... [texto completo excede {limite} caracteres]"


def _recortar_par_diff(antes: str, depois: str, contexto: int = 50) -> tuple[str, str]:
    """Recorta só a região que mudou (+ contexto), para Antes/Depois ficar legível."""
    a = (antes or "").strip()
    b = (depois or "").strip()
    if not a and not b:
        return "—", "—"
    if a == b:
        curto = _formatar_trecho(a, 120)
        return curto, curto
    if not a:
        return "—", _formatar_trecho(b, 180)
    if not b:
        return _formatar_trecho(a, 180), "—"

    sm = difflib.SequenceMatcher(None, a, b, autojunk=False)
    spans = [
        (i1, i2, j1, j2)
        for tag, i1, i2, j1, j2 in sm.get_opcodes()
        if tag != "equal"
    ]
    if not spans:
        return _formatar_trecho(a, 160), _formatar_trecho(b, 160)

    i1 = min(s[0] for s in spans)
    i2 = max(s[1] for s in spans)
    j1 = min(s[2] for s in spans)
    j2 = max(s[3] for s in spans)
    ai, aj = max(0, i1 - contexto), min(len(a), i2 + contexto)
    bi, bj = max(0, j1 - contexto), min(len(b), j2 + contexto)
    trecho_a = (("…" if ai else "") + a[ai:aj] + ("…" if aj < len(a) else ""))
    trecho_b = (("…" if bi else "") + b[bi:bj] + ("…" if bj < len(b) else ""))
    return _formatar_trecho(trecho_a, 220), _formatar_trecho(trecho_b, 220)


def _diff_so_espaco_interno(antes: str, depois: str) -> bool:
    """True se a única diferença for espaço no meio de token (ruído típico de PDF)."""
    a = re.sub(r"\s+", "", antes or "")
    b = re.sub(r"\s+", "", depois or "")
    return bool(a) and a == b and (antes or "") != (depois or "")


def _descrever_mudanca(antes: str, depois: str) -> str:
    """Frase curta em português do que mudou — o gestor precisa ler sem caça-palavras."""
    a = (antes or "").strip()
    b = (depois or "").strip()
    if not a and b:
        return f'acrescentou: "{_formatar_trecho(b, 80)}"'
    if a and not b:
        return f'removeu: "{_formatar_trecho(a, 80)}"'
    if a == b:
        return "sem diferença textual"
    if _diff_so_espaco_interno(a, b):
        return "possível ruído de leitura do PDF (só espaço no meio da palavra/número)"

    sm = difflib.SequenceMatcher(None, a, b, autojunk=False)
    partes: list[str] = []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal":
            continue
        velho = a[i1:i2]
        novo = b[j1:j2]
        if tag == "insert" and novo.strip():
            partes.append(f'acrescentou "{_formatar_trecho(novo.strip(), 60)}"')
        elif tag == "delete" and velho.strip():
            partes.append(f'removeu "{_formatar_trecho(velho.strip(), 60)}"')
        elif tag == "replace":
            if velho.strip() and novo.strip():
                partes.append(
                    f'trocou "{_formatar_trecho(velho.strip(), 40)}" '
                    f'por "{_formatar_trecho(novo.strip(), 40)}"'
                )
            elif novo.strip():
                partes.append(f'acrescentou "{_formatar_trecho(novo.strip(), 60)}"')
            elif velho.strip():
                partes.append(f'removeu "{_formatar_trecho(velho.strip(), 60)}"')
        if len(partes) >= 3:
            break
    if not partes:
        return "texto alterado (veja Antes/Depois)"
    return "; ".join(partes)


_ERROS_EXCEL_LEGIVEL = {
    "#VALUE!": "erro de fórmula (#VALOR!)",
    "#VALOR!": "erro de fórmula (#VALOR!)",
    "#REF!": "erro de referência (#REF!)",
    "#N/A": "valor indisponível (#N/D)",
    "#N/D": "valor indisponível (#N/D)",
    "#DIV/0!": "divisão por zero (#DIV/0!)",
    "#DIV/0": "divisão por zero (#DIV/0!)",
    "#NAME?": "nome inválido (#NOME?)",
    "#NOME?": "nome inválido (#NOME?)",
    "#NULL!": "interseção vazia (#NULO!)",
    "#NUM!": "número inválido (#NÚM!)",
}


_RE_CODIGO_CONTA = re.compile(r"^\d+(?:\.\d+)*$")


def _celula_ignorar_no_diff(valor: Any) -> bool:
    """Fórmulas e erros Excel não entram no Antes/Depois (geram #VALOR! sem contexto)."""
    if valor is None:
        return True
    bruto = str(valor).strip()
    if not bruto:
        return True
    if bruto.startswith("="):
        return True
    upper = bruto.upper()
    if upper in _ERROS_EXCEL_LEGIVEL or bruto.startswith("#"):
        return True
    return False


def _detectar_colunas_contas(rows: list[tuple]) -> tuple[int, int] | None:
    """Índices (conta, nome) — variações Bacen: NOME DA CONTA, DESCRIÇÃO, NOME, etc."""
    for row in rows[:25]:
        if not row:
            continue
        cols = [str(c or "").strip().upper() for c in row]
        try:
            i_conta = cols.index("CONTA")
        except ValueError:
            continue
        i_nome = next(
            (j for j, c in enumerate(cols) if j != i_conta and "NOME" in c and "CONTA" in c),
            None,
        )
        if i_nome is None:
            for j, c in enumerate(cols):
                if j == i_conta:
                    continue
                if c in ("NOME", "DESCRIÇÃO", "DESCRICAO", "DESCRIÇAO"):
                    i_nome = j
                    break
        if i_nome is None and i_conta + 1 < len(cols):
            prox = cols[i_conta + 1]
            if prox and not prox.startswith("ELEMENTO") and prox not in ("VALOR", "SALDO"):
                i_nome = i_conta + 1
        if i_nome is not None:
            return i_conta, i_nome
    return None


def _normalizar_codigo_conta(valor: Any) -> str | None:
    """Converte célula CONTA (número ou texto) para '180' / '180.01'."""
    if valor is None:
        return None
    if isinstance(valor, float):
        if valor.is_integer():
            codigo = str(int(valor))
        else:
            codigo = f"{valor:.10f}".rstrip("0").rstrip(".")
    elif isinstance(valor, int):
        codigo = str(valor)
    else:
        codigo = str(valor).strip()
    if _RE_CODIGO_CONTA.match(codigo):
        return codigo
    return None


def _extrair_mapa_contas(
    rows: list[tuple], idx_conta: int, idx_nome: int
) -> dict[str, str]:
    mapa: dict[str, str] = {}
    for row in rows:
        if not row or len(row) <= max(idx_conta, idx_nome):
            continue
        codigo_raw = row[idx_conta]
        codigo = _normalizar_codigo_conta(codigo_raw)
        if not codigo:
            continue
        nome_cel = row[idx_nome]
        if _celula_ignorar_no_diff(nome_cel):
            continue
        mapa[codigo] = _formatar_valor_planilha(nome_cel)
    return mapa


def _comparar_aba_contas_bacen(
    rows_ant: list[tuple],
    rows_atual: list[tuple],
    rotulo_aba: str,
    *,
    limite: int = 200,
) -> list[str]:
    """Diff por código contábil — ignora coluna VALOR (fórmulas)."""
    cols = _detectar_colunas_contas(rows_ant) or _detectar_colunas_contas(rows_atual)
    if not cols:
        return []
    idx_conta, idx_nome = cols
    ant = _extrair_mapa_contas(rows_ant, idx_conta, idx_nome)
    atual = _extrair_mapa_contas(rows_atual, idx_conta, idx_nome)
    alterados: list[str] = []

    def _norm(t: str) -> str:
        return re.sub(r"\s+", " ", str(t or "").strip()).upper()

    for cod in sorted(set(atual) - set(ant)):
        alterados.append(f'Aba {rotulo_aba}, conta {cod}: incluída "{atual[cod]}"')
        if len(alterados) >= limite:
            return alterados
    for cod in sorted(set(ant) - set(atual)):
        alterados.append(f'Aba {rotulo_aba}, conta {cod}: removida "{ant[cod]}"')
        if len(alterados) >= limite:
            return alterados
    for cod in sorted(set(ant) & set(atual)):
        if _norm(ant[cod]) == _norm(atual[cod]):
            continue
        alterados.append(
            f'Aba {rotulo_aba}, conta {cod}: mudanca "nome da conta"; '
            f'antes "{ant[cod]}"; depois "{atual[cod]}"'
        )
        if len(alterados) >= limite:
            return alterados
    return alterados


def _arquivo_modelo_contas_bacen(
    abas_ant: list[str],
    abas_atual: list[str],
    mapa_abas: dict[str, str],
) -> bool:
    """True se a maioria das abas parece planilha modelo documento (contas) do Bacen."""
    del abas_ant, abas_atual
    nomes = {n.lower() for n in mapa_abas.values()} | {n.lower() for n in mapa_abas.keys()}
    marcadores = {
        "flaf",
        "imob",
        "pr",
        "prcos",
        "lcsp",
        "lec",
        "loc",
        "ra",
        "rwacirb",
        "rwampad",
        "rwaopad",
        "rwasp",
        "irrbb",
    }
    return len(nomes & marcadores) >= 3


def _formatar_valor_planilha(valor: Any) -> str:
    if valor is None:
        return "em branco"
    # openpyxl pode devolver objetos Error para células com falha de fórmula.
    tipo = type(valor).__name__
    if tipo == "Error" or (
        hasattr(valor, "value") and str(getattr(valor, "value", "")).startswith("#")
    ):
        bruto = str(getattr(valor, "value", valor)).strip().upper()
    else:
        bruto = str(valor).strip()
    if bruto.startswith("="):
        return "(fórmula — ignorada no diff)"
    if bruto.upper() in _ERROS_EXCEL_LEGIVEL:
        return _ERROS_EXCEL_LEGIVEL[bruto.upper()]
    if bruto.startswith("#") and bruto.endswith("!"):
        return f"erro de planilha ({bruto})"
    if bruto.startswith("#"):
        return f"erro de planilha ({bruto})"
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y %H:%M")
    if isinstance(valor, date):
        return valor.strftime("%d/%m/%Y")
    return bruto if bruto else "em branco"


def _comparar_linhas(
    linhas_ant_num: list[tuple[int, str]],
    linhas_atual_num: list[tuple[int, str]],
    *,
    contexto: str = "",
) -> dict[str, list[str]]:
    linhas_ant = [linha for _, linha in linhas_ant_num]
    linhas_atual = [linha for _, linha in linhas_atual_num]
    prefixo = f"{contexto} - " if contexto else ""
    incluidos: list[str] = []
    removidos: list[str] = []
    alterados = []

    matcher = difflib.SequenceMatcher(None, linhas_ant, linhas_atual)
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == "insert":
            for numero, valor in linhas_atual_num[j1:j2]:
                incluidos.append(
                    f"{prefixo}linha atual {numero}: incluído \"{_formatar_trecho(valor)}\""
                )
        elif tag == "delete":
            for numero, valor in linhas_ant_num[i1:i2]:
                removidos.append(
                    f"{prefixo}linha anterior {numero}: removido \"{_formatar_trecho(valor)}\""
                )
        if tag == "replace":
            qtd = max(i2 - i1, j2 - j1)
            for idx in range(qtd):
                ant = linhas_ant_num[i1 + idx] if i1 + idx < i2 else None
                novo = linhas_atual_num[j1 + idx] if j1 + idx < j2 else None
                if ant and novo:
                    if _diff_so_espaco_interno(ant[1], novo[1]):
                        continue
                    antes_c, depois_c = _recortar_par_diff(ant[1], novo[1])
                    if antes_c.strip() == depois_c.strip():
                        continue
                    descricao = _descrever_mudanca(ant[1], novo[1])
                    if "ruído de leitura" in descricao:
                        continue
                    alterados.append(
                        f"{prefixo}linha anterior {ant[0]} -> linha atual {novo[0]}: "
                        f"mudanca \"{descricao}\"; "
                        f"antes \"{antes_c}\"; depois \"{depois_c}\""
                    )
                elif novo:
                    incluidos.append(
                        f"{prefixo}linha atual {novo[0]}: incluído \"{_formatar_trecho(novo[1])}\""
                    )
                elif ant:
                    removidos.append(
                        f"{prefixo}linha anterior {ant[0]}: removido \"{_formatar_trecho(ant[1])}\""
                    )

    return {
        "incluidos": incluidos,
        "removidos": removidos,
        "alterados": alterados,
    }


def _comparar_texto(anterior: Path, atual: Path) -> dict[str, Any]:
    diff = _comparar_linhas(
        _linhas_com_numero(_ler_texto(anterior)),
        _linhas_com_numero(_ler_texto(atual)),
    )
    incluidos = diff["incluidos"]
    removidos = diff["removidos"]
    alterados = diff["alterados"]

    return {
        "resumo_executivo": _resumo(incluidos, removidos, alterados),
        "impacto_sugerido": "Revisar os trechos alterados para avaliar impacto operacional.",
        "itens_incluidos": _limitar(incluidos),
        "itens_removidos": _limitar(removidos),
        "itens_alterados": _limitar(alterados),
    }


def _xml_tag(tag: str) -> str:
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag


def _xml_attr(elem: ElementTree.Element, nome: str) -> str:
    for chave, valor in elem.attrib.items():
        if _xml_tag(chave) == nome:
            return valor
    return ""


def _mapa_linhas_xsd(texto: str) -> dict[str, int]:
    linhas: dict[str, int] = {}
    for numero, linha in enumerate(texto.splitlines(), start=1):
        for nome in re.findall(r'\b(?:name|ref)="([^"]+)"', linha):
            linhas.setdefault(nome, numero)
    return linhas


def _coletar_xsd(path: Path) -> dict[str, dict[str, Any]]:
    texto = _ler_texto(path)
    root = ElementTree.fromstring(texto)
    linhas_por_nome = _mapa_linhas_xsd(texto)
    itens: dict[str, dict[str, Any]] = {}

    def walk(elem: ElementTree.Element, prefixo: str) -> None:
        tag = _xml_tag(elem.tag)
        nome = _xml_attr(elem, "name") or _xml_attr(elem, "ref") or tag
        atual = f"{prefixo}/{nome}" if prefixo else nome
        if tag in {"element", "attribute", "complexType", "simpleType"}:
            assinatura = {
                "tag": tag,
                "type": _xml_attr(elem, "type"),
                "base": _xml_attr(elem, "base"),
                "minOccurs": _xml_attr(elem, "minOccurs"),
                "maxOccurs": _xml_attr(elem, "maxOccurs"),
                "use": _xml_attr(elem, "use"),
            }
            assinatura_texto = ", ".join(
                f"{k}={v}" for k, v in assinatura.items() if v
            ) or tag
            itens[atual] = {
                "assinatura": assinatura_texto,
                "linha": linhas_por_nome.get(nome),
                "nome": nome,
            }
        for child in list(elem):
            walk(child, atual)

    walk(root, "")
    return itens


def _comparar_xsd(anterior: Path, atual: Path) -> dict[str, Any]:
    ant = _coletar_xsd(anterior)
    novo = _coletar_xsd(atual)
    ch_ant = set(ant)
    ch_novo = set(novo)
    incluidos = [
        f"Linha atual {novo[chave].get('linha') or '?'}: campo/schema incluído {chave} "
        f"({novo[chave]['assinatura']})"
        for chave in sorted(ch_novo - ch_ant)
    ]
    removidos = [
        f"Linha anterior {ant[chave].get('linha') or '?'}: campo/schema removido {chave} "
        f"({ant[chave]['assinatura']})"
        for chave in sorted(ch_ant - ch_novo)
    ]
    alterados = []
    for chave in sorted(ch_ant & ch_novo):
        if ant[chave]["assinatura"] != novo[chave]["assinatura"]:
            alterados.append(
                f"{chave}: linha anterior {ant[chave].get('linha') or '?'} -> "
                f"linha atual {novo[chave].get('linha') or '?'}; antes "
                f"({ant[chave]['assinatura']}); depois ({novo[chave]['assinatura']})"
            )
    return {
        "resumo_executivo": _resumo(incluidos, removidos, alterados),
        "impacto_sugerido": "Revisar campos, tipos e obrigatoriedade alterados no schema.",
        "itens_incluidos": _limitar(incluidos),
        "itens_removidos": _limitar(removidos),
        "itens_alterados": _limitar(alterados),
    }


def _classificar_mudancas_abas(
    abas_ant: list[str], abas_atual: list[str]
) -> tuple[list[str], list[str], list[str], dict[str, str]]:
    """Separa abas incluídas/removidas/renomeadas; mapa ant→atual para diff de células."""
    set_ant = set(abas_ant)
    set_atual = set(abas_atual)
    iguais = set_ant & set_atual
    ant_restante = set_ant - iguais
    atual_restante = set_atual - iguais

    ant_por_lower: dict[str, list[str]] = {}
    for nome in ant_restante:
        ant_por_lower.setdefault(nome.lower(), []).append(nome)
    atual_por_lower: dict[str, list[str]] = {}
    for nome in atual_restante:
        atual_por_lower.setdefault(nome.lower(), []).append(nome)

    renomeadas: list[str] = []
    mapa: dict[str, str] = {a: a for a in iguais}

    for chave in sorted(set(ant_por_lower) & set(atual_por_lower)):
        ants = ant_por_lower[chave]
        atuais = atual_por_lower[chave]
        if len(ants) == 1 and len(atuais) == 1 and ants[0] != atuais[0]:
            ant, novo = ants[0], atuais[0]
            renomeadas.append(f'Aba renomeada: "{ant}" → "{novo}"')
            mapa[ant] = novo
            ant_restante.discard(ant)
            atual_restante.discard(novo)

    incluidos = [f"Aba incluída: {aba}" for aba in sorted(atual_restante)]
    removidos = [f"Aba removida: {aba}" for aba in sorted(ant_restante)]
    return incluidos, removidos, renomeadas, mapa


def _comparar_xlsx(anterior: Path, atual: Path) -> dict[str, Any]:
    try:
        import openpyxl  # type: ignore
    except Exception:
        return _fallback_dependencia("XLSX", "openpyxl")

    from openpyxl.utils import get_column_letter

    # data_only=True: usa o valor que o Excel gravou ao salvar (rótulos/códigos),
    # não a fórmula bruta — evita #VALOR! e strings de fórmula no Antes/Depois.
    wb_ant = openpyxl.load_workbook(anterior, data_only=True, read_only=False)
    wb_atual = openpyxl.load_workbook(atual, data_only=True, read_only=False)

    incluidos, removidos, renomeadas, mapa_abas = _classificar_mudancas_abas(
        wb_ant.sheetnames, wb_atual.sheetnames
    )
    alterados: list[str] = list(renomeadas)
    limite_evidencias = 200

    for aba_ant, aba_atual in sorted(mapa_abas.items()):
        if len(alterados) >= limite_evidencias:
            break
        ws_ant = wb_ant[aba_ant]
        ws_atual = wb_atual[aba_atual]
        rotulo_aba = aba_atual if aba_ant == aba_atual else f"{aba_ant}→{aba_atual}"
        rows_ant = list(ws_ant.iter_rows(values_only=True))
        rows_atual = list(ws_atual.iter_rows(values_only=True))

        layout_contas = (
            _detectar_colunas_contas(rows_ant) or _detectar_colunas_contas(rows_atual)
        )
        if layout_contas:
            diff_contas = _comparar_aba_contas_bacen(
                rows_ant, rows_atual, rotulo_aba, limite=limite_evidencias
            )
            for item in diff_contas:
                if len(alterados) >= limite_evidencias:
                    break
                alterados.append(item)
            continue

        # Abas sem coluna CONTA (Capa, INSTcapital, RWAopad): diff célula a célula
        # desalinha linhas e polui Antes/Depois — ignorar no modelo documento (contas).
        if _arquivo_modelo_contas_bacen(wb_ant.sheetnames, wb_atual.sheetnames, mapa_abas):
            continue

        cabecalhos = rows_atual[0] if rows_atual else ()
        max_row = max(len(rows_ant), len(rows_atual))

        for row_idx in range(max_row):
            if len(alterados) >= limite_evidencias:
                break
            row_ant = rows_ant[row_idx] if row_idx < len(rows_ant) else ()
            row_atual = rows_atual[row_idx] if row_idx < len(rows_atual) else ()
            max_col = max(len(row_ant), len(row_atual))
            for col_idx in range(max_col):
                v_ant = row_ant[col_idx] if col_idx < len(row_ant) else None
                v_atual = row_atual[col_idx] if col_idx < len(row_atual) else None
                if _celula_ignorar_no_diff(v_ant) and _celula_ignorar_no_diff(v_atual):
                    continue
                if _celula_ignorar_no_diff(v_ant) or _celula_ignorar_no_diff(v_atual):
                    continue
                v_ant_fmt = _formatar_valor_planilha(v_ant)
                v_atual_fmt = _formatar_valor_planilha(v_atual)
                if v_ant_fmt == v_atual_fmt:
                    continue
                coluna = get_column_letter(col_idx + 1)
                cabecalho = cabecalhos[col_idx] if col_idx < len(cabecalhos) else None
                contexto = f", coluna {cabecalho}" if cabecalho and row_idx != 0 else ""
                alterados.append(
                    f"Aba {rotulo_aba}, célula {coluna}{row_idx + 1}{contexto}: "
                    f'antes "{v_ant_fmt}"; '
                    f'depois "{v_atual_fmt}"'
                )
                if len(alterados) >= limite_evidencias:
                    break

    wb_ant.close()
    wb_atual.close()

    return {
        "resumo_executivo": _resumo(incluidos, removidos, alterados),
        "impacto_sugerido": "Revisar abas e células alteradas antes de atualizar rotinas internas.",
        "itens_incluidos": _limitar(incluidos),
        "itens_removidos": _limitar(removidos),
        "itens_alterados": _limitar(alterados, 500),
    }


def _coluna_excel(col_zero_based: int) -> str:
    col = col_zero_based + 1
    letras = ""
    while col:
        col, resto = divmod(col - 1, 26)
        letras = chr(65 + resto) + letras
    return letras


def _comparar_xls(anterior: Path, atual: Path) -> dict[str, Any]:
    try:
        import xlrd  # type: ignore
    except Exception:
        return _fallback_dependencia("XLS", "xlrd")

    wb_ant = xlrd.open_workbook(str(anterior))
    wb_atual = xlrd.open_workbook(str(atual))
    incluidos, removidos, renomeadas, mapa_abas = _classificar_mudancas_abas(
        wb_ant.sheet_names(), wb_atual.sheet_names()
    )
    alterados: list[str] = list(renomeadas)

    for aba_ant, aba_atual in sorted(mapa_abas.items()):
        sh_ant = wb_ant.sheet_by_name(aba_ant)
        sh_atual = wb_atual.sheet_by_name(aba_atual)
        rotulo_aba = aba_atual if aba_ant == aba_atual else f"{aba_ant}→{aba_atual}"
        max_row = max(sh_ant.nrows, sh_atual.nrows)
        max_col = max(sh_ant.ncols, sh_atual.ncols)
        for row in range(max_row):
            for col in range(max_col):
                v_ant = sh_ant.cell_value(row, col) if row < sh_ant.nrows and col < sh_ant.ncols else None
                v_atual = sh_atual.cell_value(row, col) if row < sh_atual.nrows and col < sh_atual.ncols else None
                if v_ant != v_atual:
                    coluna = _coluna_excel(col)
                    cabecalho = (
                        sh_atual.cell_value(0, col)
                        if row != 0 and sh_atual.nrows and col < sh_atual.ncols
                        else ""
                    )
                    contexto = f", coluna {cabecalho}" if cabecalho else ""
                    alterados.append(
                        f"Aba {rotulo_aba}, célula {coluna}{row + 1}{contexto}: "
                        f'antes "{_formatar_valor_planilha(v_ant)}"; '
                        f'depois "{_formatar_valor_planilha(v_atual)}"'
                    )
                    if len(alterados) >= 200:
                        break
            if len(alterados) >= 200:
                break

    return {
        "resumo_executivo": _resumo(incluidos, removidos, alterados),
        "impacto_sugerido": "Revisar abas e células alteradas antes de atualizar rotinas internas.",
        "itens_incluidos": _limitar(incluidos),
        "itens_removidos": _limitar(removidos),
        "itens_alterados": _limitar(alterados),
    }


def _comparar_pdf(anterior: Path, atual: Path) -> dict[str, Any]:
    try:
        from pypdf import PdfReader  # type: ignore
    except Exception:
        return _fallback_dependencia("PDF", "pypdf")

    def extrair_paginas(path: Path) -> list[list[tuple[int, str]]]:
        reader = PdfReader(str(path))
        paginas = []
        for page in reader.pages:
            paginas.append(_linhas_com_numero(page.extract_text() or ""))
        return paginas

    # Ruído típico de cabeçalho/rodapé em PDF do Bacen — não ajuda o gestor.
    ruido = {"interno", "confidencial", "página", "pagina"}

    paginas_ant = extrair_paginas(anterior)
    paginas_atual = extrair_paginas(atual)
    incluidos: list[str] = []
    removidos: list[str] = []
    alterados: list[str] = []
    total_paginas = max(len(paginas_ant), len(paginas_atual))
    for idx in range(total_paginas):
        ant = paginas_ant[idx] if idx < len(paginas_ant) else []
        novo = paginas_atual[idx] if idx < len(paginas_atual) else []
        diff = _comparar_linhas(ant, novo, contexto=f"Página {idx + 1}")
        for item in diff["incluidos"]:
            trecho = item.rsplit(': incluído "', 1)[-1].rstrip('"').strip().lower()
            if trecho in ruido or len(trecho) <= 2:
                continue
            incluidos.append(item)
        for item in diff["removidos"]:
            trecho = item.rsplit(': removido "', 1)[-1].rstrip('"').strip().lower()
            if trecho in ruido or len(trecho) <= 2:
                continue
            removidos.append(item)
        for item in diff["alterados"]:
            # Descarta "mudança" em que o recorte ficou idêntico (só ruído de layout).
            if 'antes "' in item and 'depois "' in item:
                try:
                    if 'mudanca "' in item:
                        meio = item.split('"; antes "', 1)[1]
                    else:
                        meio = item.split(': antes "', 1)[1]
                    ant_t, dep_t = meio.rsplit('"; depois "', 1)
                    dep_t = dep_t.rstrip('"')
                    if ant_t.strip() == dep_t.strip():
                        continue
                    if _diff_so_espaco_interno(ant_t, dep_t):
                        continue
                except Exception:
                    pass
            alterados.append(item)

    return {
        "resumo_executivo": _resumo(incluidos, removidos, alterados),
        "impacto_sugerido": "Revisar as páginas e linhas alteradas no PDF.",
        "itens_incluidos": _limitar(incluidos),
        "itens_removidos": _limitar(removidos),
        "itens_alterados": _limitar(alterados),
    }


def _comparar_zip(anterior: Path, atual: Path) -> dict[str, Any]:
    with zipfile.ZipFile(anterior) as z_ant, zipfile.ZipFile(atual) as z_atual:
        ant_infos = {i.filename: i for i in z_ant.infolist()}
        novo_infos = {i.filename: i for i in z_atual.infolist()}
        ant = set(ant_infos)
        novo = set(novo_infos)
        incluidos = []
        removidos = []
        alterados = []

        for nome in sorted(novo - ant):
            data = z_atual.read(nome)
            evidencia = _evidencia_zip_texto(data)
            incluidos.append(
                f"Arquivo interno incluído: {nome} ({len(data)} bytes){evidencia}"
            )
        for nome in sorted(ant - novo):
            data = z_ant.read(nome)
            evidencia = _evidencia_zip_texto(data)
            removidos.append(
                f"Arquivo interno removido: {nome} ({len(data)} bytes){evidencia}"
            )
        for nome in sorted(ant & novo):
            data_ant = z_ant.read(nome)
            data_novo = z_atual.read(nome)
            if hashlib.sha256(data_ant).hexdigest() != hashlib.sha256(data_novo).hexdigest():
                if _parece_texto(nome):
                    diff = _comparar_linhas(
                        _linhas_com_numero(_bytes_para_texto(data_ant)),
                        _linhas_com_numero(_bytes_para_texto(data_novo)),
                        contexto=f"Arquivo interno {nome}",
                    )
                    alterados.extend(diff["alterados"])
                    incluidos.extend(diff["incluidos"])
                    removidos.extend(diff["removidos"])
                else:
                    alterados.append(
                        f"Arquivo interno {nome}: conteúdo binário alterado; "
                        f"tamanho anterior {len(data_ant)} bytes; tamanho atual {len(data_novo)} bytes"
                    )
    return {
        "resumo_executivo": _resumo(incluidos, removidos, alterados),
        "impacto_sugerido": "Revisar arquivos internos alterados no pacote ZIP.",
        "itens_incluidos": _limitar(incluidos),
        "itens_removidos": _limitar(removidos),
        "itens_alterados": _limitar(alterados),
    }


def _bytes_para_texto(data: bytes) -> str:
    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="ignore")


def _parece_texto(nome: str) -> bool:
    return Path(nome).suffix.lower() in {
        ".txt",
        ".csv",
        ".xml",
        ".xsd",
        ".json",
        ".html",
        ".htm",
    }


def _evidencia_zip_texto(data: bytes) -> str:
    texto = _bytes_para_texto(data)
    linhas = _normalizar_linhas(texto)
    if not linhas:
        return ""
    return f"; evidência: \"{_formatar_trecho(linhas[0], 180)}\""


def _fallback_dependencia(tipo: str, pacote: str) -> dict[str, Any]:
    return {
        "resumo_executivo": (
            f"Arquivo {tipo} alterado. A extração detalhada requer a dependência {pacote}."
        ),
        "impacto_sugerido": f"Instalar {pacote} para comparar conteúdo {tipo} em detalhe.",
        "itens_incluidos": [],
        "itens_removidos": [],
        "itens_alterados": [f"Comparação detalhada pendente: dependência {pacote} ausente."],
    }


def _resumo(incluidos: list[str], removidos: list[str], alterados: list[str]) -> str:
    total = len(incluidos) + len(removidos) + len(alterados)
    if total == 0:
        return "Nenhuma diferença textual/estrutural relevante foi identificada."
    return (
        f"Foram identificadas {len(incluidos)} inclusão(ões), "
        f"{len(removidos)} remoção(ões) e {len(alterados)} alteração(ões)."
    )


def comparar_arquivos(
    *,
    caminho_anterior: str | None,
    caminho_atual: str | None,
    tipo_arquivo: str,
) -> dict[str, Any] | None:
    anterior = _resolver(caminho_anterior)
    atual = _resolver(caminho_atual)
    if not anterior or not atual or not anterior.exists() or not atual.exists():
        return None

    tipo = tipo_arquivo.lower().lstrip(".")
    try:
        if tipo == "xsd":
            return _comparar_xsd(anterior, atual)
        if tipo in {"xml", "txt", "csv", "json", "html", "htm"}:
            return _comparar_texto(anterior, atual)
        if tipo in {"xlsx", "xlsm"}:
            return _comparar_xlsx(anterior, atual)
        if tipo == "xls":
            return _comparar_xls(anterior, atual)
        if tipo == "pdf":
            return _comparar_pdf(anterior, atual)
        if tipo == "zip":
            return _comparar_zip(anterior, atual)
    except Exception as exc:
        return {
            "resumo_executivo": f"Falha ao comparar conteúdo do arquivo: {exc}",
            "impacto_sugerido": "Revisar manualmente o arquivo alterado.",
            "itens_incluidos": [],
            "itens_removidos": [],
            "itens_alterados": [str(exc)],
        }
    return _comparar_texto(anterior, atual)
