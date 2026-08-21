# -*- coding: utf-8 -*-
"""Planilha única do comunicado — linguagem simples (e-mail e Exportar)."""
from __future__ import annotations

import math
from dataclasses import dataclass, field
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

BLUE = "2E3192"
BLUE_SOFT = "EEF0FA"
TEXT = "1F2937"
MUTED = "64748B"
GRAY = "F8F9FA"
GREEN_SOFT = "E8F7EE"
YELLOW_SOFT = "FFF8E6"
ROW_ALT = "F3F4F8"
WHITE = "FFFFFF"
BORDER = "E5E7EB"
LINK = "1D4ED8"
ROTULO_LINK = "Abrir no site do Bacen"

# Linhas fixas em todas as abas
ROW_TITULO = 1
ROW_EXPLICACAO = 2
ROW_CABECALHO = 4
ROW_DADOS = 5


@dataclass
class ArquivoResumoPlanilha:
    data: str
    leiaute: str
    arquivo: str
    situacao: str
    precisa_agir: bool
    qtd_mudancas: int
    link: str
    observacao: str = ""


@dataclass
class LinhaMudancaPlanilha:
    data: str
    leiaute: str
    arquivo: str
    onde: str
    o_que_mudou: str
    antes: str
    depois: str
    o_que_fazer: str


@dataclass
class DadosPlanilhaGestor:
    arquivos_agir: list[ArquivoResumoPlanilha] = field(default_factory=list)
    linhas_mudanca: list[LinhaMudancaPlanilha] = field(default_factory=list)
    arquivos_aviso: list[ArquivoResumoPlanilha] = field(default_factory=list)


def _borda() -> Border:
    lado = Side(style="thin", color=BORDER)
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _escrever_topo(ws: Worksheet, n_cols: int, titulo: str, explicacao: str) -> None:
    ultima = get_column_letter(n_cols)
    ws.merge_cells(f"A{ROW_TITULO}:{ultima}{ROW_TITULO}")
    ws.merge_cells(f"A{ROW_EXPLICACAO}:{ultima}{ROW_EXPLICACAO}")

    c1 = ws.cell(row=ROW_TITULO, column=1, value=titulo)
    c1.font = Font(name="Arial", size=16, bold=True, color=BLUE)
    c1.alignment = Alignment(vertical="center", wrap_text=True)
    c1.fill = PatternFill("solid", fgColor=BLUE_SOFT)
    ws.row_dimensions[ROW_TITULO].height = 30

    c2 = ws.cell(row=ROW_EXPLICACAO, column=1, value=explicacao)
    c2.font = Font(name="Arial", size=10, color=MUTED)
    c2.alignment = Alignment(vertical="center", wrap_text=True)
    c2.fill = PatternFill("solid", fgColor=GRAY)
    # Largura aproximada da faixa mesclada para calcular altura da explicação.
    largura_faixa = max(40.0, 12.0 * n_cols)
    linhas_exp = _contar_linhas_visuais(explicacao, largura_faixa)
    ws.row_dimensions[ROW_EXPLICACAO].height = max(36, min(72, 10 + linhas_exp * 14))

    ws.row_dimensions[3].height = 10


def _escrever_cabecalho(ws: Worksheet, colunas: list[str]) -> None:
    for col, nome in enumerate(colunas, start=1):
        cell = ws.cell(row=ROW_CABECALHO, column=col, value=nome)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = _borda()
    ws.row_dimensions[ROW_CABECALHO].height = 38


def _estilo_dado(cell, *, horizontal: str = "left", fill: str | None = None) -> None:
    cell.font = Font(name="Arial", size=10, color=TEXT)
    cell.alignment = Alignment(vertical="top", horizontal=horizontal, wrap_text=True)
    cell.border = _borda()
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)


def _link(cell, url: str | None) -> None:
    """Mostra texto curto clicável; a URL fica só no hiperlink (não quebra feio)."""
    destino = str(url or "").strip()
    if not destino:
        cell.value = ""
        return
    cell.value = ROTULO_LINK
    cell.hyperlink = destino
    cell.font = Font(name="Arial", size=10, color=LINK, underline="single")
    cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
    cell.border = _borda()


def _larguras(ws: Worksheet, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _contar_linhas_visuais(texto: object, largura_col: float) -> int:
    """Estima quantas linhas o Excel precisa com wrap (openpyxl não autoajusta sozinho)."""
    bruto = str(texto if texto is not None else "")
    if not bruto:
        return 1
    # Largura da coluna no Excel ≈ caracteres da fonte Arial 10.
    chars_por_linha = max(6, int(largura_col * 0.92))
    total = 0
    for trecho in bruto.replace("\r\n", "\n").split("\n"):
        if trecho == "":
            total += 1
            continue
        total += max(1, math.ceil(len(trecho) / chars_por_linha))
    return max(1, total)


def _ajustar_altura_linhas(
    ws: Worksheet,
    widths: dict[str, float],
    *,
    n_cols: int,
    primeira: int = ROW_DADOS,
    min_altura: float = 30,
    max_altura: float = 140,
    px_por_linha: float = 15,
) -> None:
    for row_idx in range(primeira, ws.max_row + 1):
        max_linhas = 1
        for col in range(1, n_cols + 1):
            letra = get_column_letter(col)
            largura = widths.get(letra, 12.0)
            valor = ws.cell(row=row_idx, column=col).value
            max_linhas = max(max_linhas, _contar_linhas_visuais(valor, largura))
        altura = min(max_altura, max(min_altura, 8 + max_linhas * px_por_linha))
        ws.row_dimensions[row_idx].height = altura


def _fechar_aba(ws: Worksheet, n_cols: int, widths: dict[str, float]) -> None:
    ultima = get_column_letter(n_cols)
    fim = max(ROW_CABECALHO, ws.max_row)
    ws.auto_filter.ref = f"A{ROW_CABECALHO}:{ultima}{fim}"
    ws.freeze_panes = f"A{ROW_DADOS}"
    ws.sheet_view.showGridLines = False
    _ajustar_altura_linhas(ws, widths, n_cols=n_cols)


def _escrever_linha(ws: Worksheet, row: int, valores: list) -> None:
    for col, valor in enumerate(valores, start=1):
        ws.cell(row=row, column=col, value=valor)


def gerar_bytes_planilha_gestor(
    dados: DadosPlanilhaGestor,
    *,
    nome_arquivo: str,
) -> tuple[bytes, str]:
    """
    Três abas com título explicativo e formatação legível:
    Resumo · O que mudou · Só aviso
    """
    wb = Workbook()

    # --- Resumo ---
    ws = wb.active
    ws.title = "Resumo"
    cols = [
        "Data",
        "Leiaute",
        "Arquivo",
        "Situação",
        "Precisa agir?",
        "Quantidade de mudanças",
        "Link Bacen",
    ]
    _escrever_topo(
        ws,
        len(cols),
        "Resumo — visão geral dos arquivos",
        (
            "Aqui você vê um arquivo por linha. "
            "Use a coluna “Precisa agir?” para saber o que revisar. "
            "“Sim” = houve mudança de conteúdo. “Não” = o Bacen só republicou no site."
        ),
    )
    _escrever_cabecalho(ws, cols)

    row = ROW_DADOS
    linhas_resumo: list[list] = []
    for arq in dados.arquivos_agir:
        linhas_resumo.append(
            [arq.data, arq.leiaute, arq.arquivo, arq.situacao, "Sim", arq.qtd_mudancas, arq.link]
        )
    for arq in dados.arquivos_aviso:
        linhas_resumo.append(
            [arq.data, arq.leiaute, arq.arquivo, arq.situacao, "Não", 0, arq.link]
        )
    if not linhas_resumo:
        linhas_resumo.append(
            ["", "", "", "Nenhuma mudança nesta exportação.", "", "", ""]
        )

    for i, valores in enumerate(linhas_resumo):
        url = valores[6]
        valores_visuais = list(valores)
        valores_visuais[6] = ROTULO_LINK if url else ""
        _escrever_linha(ws, row, valores_visuais)
        zebra = ROW_ALT if i % 2 == 1 else None
        for col in range(1, len(cols) + 1):
            cell = ws.cell(row=row, column=col)
            fill = zebra
            horizontal = "left"
            if col == 5:
                horizontal = "center"
                if cell.value == "Sim":
                    fill = GREEN_SOFT
                elif cell.value == "Não":
                    fill = YELLOW_SOFT
            elif col == 6:
                horizontal = "center"
            _estilo_dado(cell, horizontal=horizontal, fill=fill)
        _link(ws.cell(row=row, column=7), url if url else None)
        row += 1

    widths_resumo = {"A": 18, "B": 12, "C": 34, "D": 28, "E": 14, "F": 14, "G": 24}
    _larguras(ws, widths_resumo)
    _fechar_aba(ws, len(cols), widths_resumo)

    # --- O que mudou ---
    ws = wb.create_sheet("O que mudou")
    cols = [
        "Data",
        "Leiaute",
        "Arquivo",
        "Onde",
        "O que mudou",
        "Antes",
        "Depois",
        "O que fazer",
    ]
    _escrever_topo(
        ws,
        len(cols),
        "O que mudou — detalhe das diferenças",
        (
            "Cada linha é uma mudança. "
            "Veja onde está (aba, célula ou página), o que mudou, o valor de antes e o de depois, "
            "e o que fazer na rotina."
        ),
    )
    _escrever_cabecalho(ws, cols)

    row = ROW_DADOS
    linhas = list(dados.linhas_mudanca)
    if not linhas:
        _escrever_linha(
            ws,
            row,
            ["", "", "", "", "Nenhuma diferença detalhada nesta exportação.", "", "", ""],
        )
        for col in range(1, len(cols) + 1):
            _estilo_dado(ws.cell(row=row, column=col))
    else:
        for i, lin in enumerate(linhas):
            _escrever_linha(
                ws,
                row,
                [
                    lin.data,
                    lin.leiaute,
                    lin.arquivo,
                    lin.onde,
                    lin.o_que_mudou,
                    lin.antes,
                    lin.depois,
                    lin.o_que_fazer,
                ],
            )
            fill = ROW_ALT if i % 2 == 1 else None
            for col in range(1, len(cols) + 1):
                _estilo_dado(ws.cell(row=row, column=col), fill=fill)
            row += 1

    widths_det = {
        "A": 16,
        "B": 12,
        "C": 28,
        "D": 22,
        "E": 22,
        "F": 32,
        "G": 32,
        "H": 36,
    }
    _larguras(ws, widths_det)
    _fechar_aba(ws, len(cols), widths_det)

    # --- Só aviso ---
    ws = wb.create_sheet("Só aviso")
    cols = ["Data", "Leiaute", "Arquivo", "Observação", "Link Bacen"]
    _escrever_topo(
        ws,
        len(cols),
        "Só aviso — sem ação de conteúdo",
        (
            "Nestes arquivos o Bacen republicou no site, mas o texto, a célula ou a tabela "
            "não mudaram. Em geral não exige ajuste de rotina — só fique ciente."
        ),
    )
    _escrever_cabecalho(ws, cols)

    row = ROW_DADOS
    avisos = list(dados.arquivos_aviso)
    if not avisos:
        _escrever_linha(
            ws,
            row,
            ["", "", "", "Nenhum arquivo só de aviso nesta exportação.", ""],
        )
        for col in range(1, len(cols) + 1):
            _estilo_dado(ws.cell(row=row, column=col))
    else:
        for i, arq in enumerate(avisos):
            _escrever_linha(
                ws,
                row,
                [
                    arq.data,
                    arq.leiaute,
                    arq.arquivo,
                    arq.observacao
                    or "O Bacen republicou o arquivo no site, sem mudança de conteúdo.",
                    ROTULO_LINK if arq.link else "",
                ],
            )
            fill = ROW_ALT if i % 2 == 1 else None
            for col in range(1, 5):
                _estilo_dado(ws.cell(row=row, column=col), fill=fill)
            _estilo_dado(ws.cell(row=row, column=5), fill=fill)
            _link(ws.cell(row=row, column=5), arq.link)
            row += 1

    widths_aviso = {"A": 18, "B": 12, "C": 34, "D": 55, "E": 24}
    _larguras(ws, widths_aviso)
    _fechar_aba(ws, len(cols), widths_aviso)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), nome_arquivo


def rotulo_situacao(codigo: str) -> str:
    mapa = {
        "versao_pareada": "Arquivo novo na página",
        "sem_anterior": "Arquivo novo sem versão anterior",
        "mesmo_arquivo": "Mesmo arquivo atualizado",
        "aviso": "Republicado no site (sem mudança de conteúdo)",
    }
    return mapa.get(codigo, "Arquivo atualizado")
