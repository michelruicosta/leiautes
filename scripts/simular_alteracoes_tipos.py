# -*- coding: utf-8 -*-
"""Gera alteracoes simuladas para validar o comparador nas telas.

Uso:
    python scripts/simular_alteracoes_tipos.py

O script cria pares de arquivos em storage/simulacoes e registra uma execucao
com diferencas para XSD, XML, PDF, XLSX, ZIP e TXT.
"""
from __future__ import annotations

import hashlib
import sys
import zipfile
from datetime import datetime
from pathlib import Path

BASE = Path(__file__).resolve().parent.parent
if str(BASE) not in sys.path:
    sys.path.insert(0, str(BASE))

from persistencia.arquivos_db import registrar_arquivo_observado
from persistencia.db import finalizar_execucao, iniciar_execucao


SIM_DIR = BASE / "storage" / "simulacoes" / datetime.now().strftime("%Y%m%d_%H%M%S")


def _write(path: Path, data: bytes) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(data)
    return path


def _info(path: Path, etag: str) -> dict:
    data = path.read_bytes()
    return {
        "etag": etag,
        "last_modified": datetime.now().strftime("%a, %d %b %Y %H:%M:%S GMT"),
        "content_length": str(len(data)),
        "final_url": f"https://simulacao.local/{path.name}",
        "partial_fp": hashlib.sha256(data[:1024]).hexdigest(),
        "checked_at": datetime.now().isoformat(timespec="seconds"),
    }


def _registrar_par(
    *,
    execucao_id: int,
    tipo: str,
    nome: str,
    anterior: Path,
    atual: Path,
    evidencia: str,
) -> None:
    url = f"https://simulacao.local/{tipo}/{nome}"
    registrar_arquivo_observado(
        url=url,
        nome_arquivo=nome,
        info=_info(anterior, f"sim-{tipo}-v1"),
        categoria="SCD - 4111",
        execucao_id=None,
        mudou=True,
        evidencia="baseline simulada",
        caminho_arquivo=str(anterior),
    )
    registrar_arquivo_observado(
        url=url,
        nome_arquivo=nome,
        info=_info(atual, f"sim-{tipo}-v2"),
        categoria="SCD - 4111",
        execucao_id=execucao_id,
        mudou=True,
        evidencia=evidencia,
        caminho_arquivo=str(atual),
    )


def _pdf(path: Path, linhas: list[str]) -> Path:
    try:
        from reportlab.pdfgen import canvas
    except Exception as exc:
        raise RuntimeError(
            "A simulacao de PDF requer reportlab instalado localmente."
        ) from exc

    path.parent.mkdir(parents=True, exist_ok=True)
    c = canvas.Canvas(str(path))
    y = 740
    for linha in linhas:
        c.drawString(72, y, linha)
        y -= 20
    c.save()
    return path


def _xlsx(path: Path, valores: list[tuple[str, str]]) -> Path:
    from openpyxl import Workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Campos"
    ws["A1"] = "Campo"
    ws["B1"] = "Obrigatorio"
    for idx, (campo, obrigatorio) in enumerate(valores, start=2):
        ws.cell(row=idx, column=1, value=campo)
        ws.cell(row=idx, column=2, value=obrigatorio)
    wb.save(path)
    return path


def _zip(path: Path, arquivos: dict[str, bytes]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        for nome, data in arquivos.items():
            zf.writestr(nome, data)
    return path


def main() -> None:
    execucao_id = iniciar_execucao(log_path="simulacao")

    xsd_a = _write(
        SIM_DIR / "xsd" / "anterior.xsd",
        b"""<xs:schema xmlns:xs="http://www.w3.org/2001/XMLSchema">
  <xs:element name="IdentificadorRemessa" type="xs:string"/>
</xs:schema>
""",
    )
    xsd_b = _write(
        SIM_DIR / "xsd" / "atual.xsd",
        b"""<xs:schema xmlns:xs="http://www.w3.org/2001/XMLSchema">
  <xs:element name="IdentificadorRemessa" type="xs:int"/>
  <xs:element name="CodigoControle" type="xs:string"/>
</xs:schema>
""",
    )
    _registrar_par(
        execucao_id=execucao_id,
        tipo="xsd",
        nome="simulacao_schema.xsd",
        anterior=xsd_a,
        atual=xsd_b,
        evidencia="campo incluido e tipo alterado",
    )

    xml_a = _write(
        SIM_DIR / "xml" / "anterior.xml",
        b"""<root>
  <prazo>D+1 10h</prazo>
  <status>ativo</status>
</root>
""",
    )
    xml_b = _write(
        SIM_DIR / "xml" / "atual.xml",
        b"""<root>
  <prazo>D+1 09h30</prazo>
  <status>ativo</status>
  <controle>CodigoControle obrigatorio</controle>
</root>
""",
    )
    _registrar_par(
        execucao_id=execucao_id,
        tipo="xml",
        nome="simulacao_xml.xml",
        anterior=xml_a,
        atual=xml_b,
        evidencia="prazo alterado e elemento incluido",
    )

    pdf_a = _pdf(
        SIM_DIR / "pdf" / "anterior.pdf",
        ["Leiaute SCD 4111", "Prazo D+1 ate 10h", "Campo controle opcional"],
    )
    pdf_b = _pdf(
        SIM_DIR / "pdf" / "atual.pdf",
        ["Leiaute SCD 4111", "Prazo D+1 ate 09h30", "Campo controle obrigatorio"],
    )
    _registrar_par(
        execucao_id=execucao_id,
        tipo="pdf",
        nome="simulacao_pdf.pdf",
        anterior=pdf_a,
        atual=pdf_b,
        evidencia="texto de prazo e obrigatoriedade alterados",
    )

    xlsx_a = _xlsx(
        SIM_DIR / "xlsx" / "anterior.xlsx",
        [("IdentificadorRemessa", "Nao"), ("CodigoInstituicao", "Sim")],
    )
    xlsx_b = _xlsx(
        SIM_DIR / "xlsx" / "atual.xlsx",
        [("IdentificadorRemessa", "Sim"), ("CodigoInstituicao", "Sim")],
    )
    _registrar_par(
        execucao_id=execucao_id,
        tipo="xlsx",
        nome="simulacao_planilha.xlsx",
        anterior=xlsx_a,
        atual=xlsx_b,
        evidencia="celula de obrigatoriedade alterada",
    )

    zip_a = _zip(
        SIM_DIR / "zip" / "anterior.zip",
        {
            "schema.xsd": b'<xs:element name="IdentificadorRemessa" type="xs:string"/>',
            "manual.txt": b"Prazo de envio: ate 10h",
        },
    )
    zip_b = _zip(
        SIM_DIR / "zip" / "atual.zip",
        {
            "schema.xsd": b'<xs:element name="IdentificadorRemessa" type="xs:int"/>',
            "manual.txt": b"Prazo de envio: ate 09h30",
            "novo.txt": b"Novo controle: CodigoControle obrigatorio",
        },
    )
    _registrar_par(
        execucao_id=execucao_id,
        tipo="zip",
        nome="simulacao_pacote.zip",
        anterior=zip_a,
        atual=zip_b,
        evidencia="arquivo interno incluido e schema alterado",
    )

    txt_a = _write(
        SIM_DIR / "txt" / "anterior.txt",
        b"Regra de envio: remessa diaria\nPrazo de envio: ate 10h do dia util seguinte\nResponsavel: Contabilidade\n",
    )
    txt_b = _write(
        SIM_DIR / "txt" / "atual.txt",
        b"Regra de envio: remessa diaria\nPrazo de envio: ate 09h30 do dia util seguinte\nResponsavel: Contabilidade\nNovo campo obrigatorio: CodigoControle\n",
    )
    _registrar_par(
        execucao_id=execucao_id,
        tipo="txt",
        nome="simulacao_texto.txt",
        anterior=txt_a,
        atual=txt_b,
        evidencia="linha alterada e linha incluida",
    )

    finalizar_execucao(
        execucao_id,
        status="sucesso",
        qtd_leiautes=1,
        qtd_arquivos=6,
        qtd_alteracoes=6,
    )
    print(f"Simulacao registrada na execucao {execucao_id}")


if __name__ == "__main__":
    main()
