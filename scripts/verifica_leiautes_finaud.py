# -*- coding: utf-8 -*-
"""
Monitor de leiautes Bacen (v3.2a+4111)
- Logo inline por CID
- Config de e-mail em JSON (config_email.json no diretório do projeto)
- Data "hoje" dinâmica + MONITOR_TEST_DATE opcional
- Playwright com flags para ambiente compartilhado
- **NOVO**: Envia e-mail mesmo sem novidades (configurável) e deixa o texto de "Não há documentos" alinhado à esquerda e na cor azul do logotipo (#2e3192)
- **NOVO**: Suporte ao Documento 4111 (Saldos Contábeis Diários - SCD)
"""

from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path
from playwright.sync_api import sync_playwright
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
from email.mime.base import MIMEBase
from email.utils import make_msgid
from email import encoders

import html, ssl, smtplib, os, re, json, hashlib, requests, sys, mimetypes, traceback, difflib
from urllib.parse import urlparse, unquote

# >>> ajuste este caminho por projeto
TAIL_PATH_BASE = "/home/tsalachtech.com.br/public_html/monitoramentos/leiautes/_status_tail.txt"
LOG_PATH_BASE  = "/home/tsalachtech.com.br/apps/leiautes/logs/execucao_{data}.log"


def _write_status_tail(proj: str,
                       header_status: str,
                       resumo: dict,
                       ultimos: list[str] | None = None,
                       extra_info: str | None = None) -> None:
    if os.environ.get("LEIAUTES_DISABLE_STATUS_TAIL", "").strip().lower() in (
        "1",
        "true",
        "yes",
    ):
        logger.info("Status tail legado desativado por LEIAUTES_DISABLE_STATUS_TAIL=1.")
        return

    status_code = 0
    if "AVISO" in header_status:
        status_code = 1
    elif "ERRO" in header_status:
        status_code = 2

    now = datetime.now()
    now_fmt = now.strftime('%d/%m/%Y %H:%M:%S')
    tail_path = TAIL_PATH_BASE.format(proj=proj)
    log_path  = LOG_PATH_BASE.format(proj=proj, data=now.strftime("%Y%m%d"))

    lines = []
    lines.append(f"{now_fmt} | {header_status}")
    if extra_info:
        lines.append(extra_info)

    lines.append("")
    lines.append("📊 Resumo:")
    for k, v in resumo.items():
        lines.append(f"- {k}: {v}")

    if ultimos:
        lines.append("")
        lines.append("📜 Últimos itens lidos:")
        for t in ultimos[:3]:
            lines.append(f"- {t[:120]}")

    lines.append("")
    lines.append(f"ℹ️ Log completo: {log_path}")

    bloco = []
    bloco.append(f"===== INÍCIO {now_fmt} =====")
    bloco.extend(lines)
    bloco.append(f"===== FIM {now_fmt} =====")

    with open(tail_path, "w", encoding="utf-8") as f:
        f.write("\n".join(bloco) + "\n\n")


# ====== CAMINHOS/TEMPOS ======
SCRIPT_DIR = Path(__file__).resolve().parent
BASE = SCRIPT_DIR.parent
if str(BASE) not in sys.path:
    sys.path.insert(0, str(BASE))
CONFIG_PATH = BASE / "config" / "config_email.json"
LOGO_PATH = BASE / "logotipo" / "FINAUD_TEC_LOG.jpg"

try:
    from persistencia.arquivos_db import (
        registrar_arquivo_observado,
        salvar_conteudo_versao,
    )
    from persistencia.db import (
        conectar,
        finalizar_execucao,
        init_db,
        iniciar_execucao,
    )
    from persistencia.execucoes_db import contar_resultados_execucao
except Exception:
    registrar_arquivo_observado = None
    salvar_conteudo_versao = None
    conectar = None
    init_db = None
    iniciar_execucao = None
    finalizar_execucao = None
    contar_resultados_execucao = None

CONNECT_TIMEOUT = 10
READ_TIMEOUT = 10
TIMEOUT = (CONNECT_TIMEOUT, READ_TIMEOUT)

# ====== LOGGING ======
import logging
from logging.handlers import RotatingFileHandler

LOG_DIR = BASE / "logs"
LOG_DIR.mkdir(exist_ok=True)
LOG_FILE = LOG_DIR / f"monitor_leiautes_{datetime.now():%Y%m%d}.log"

logger = logging.getLogger("monitor_leiautes")
logger.setLevel(logging.INFO)
if not logger.handlers:
    fh = RotatingFileHandler(LOG_FILE, maxBytes=1_000_000, backupCount=7, encoding="utf-8")
    fh.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    sh = logging.StreamHandler(sys.stdout)
    sh.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    logger.addHandler(fh)
    logger.addHandler(sh)

# ====== AJUSTES (fallback; preferir configuracoes do banco) ======
QUIET_BASELINE = True
ONLY_ATUAL = True
EXCLUDE_PATTERNS = ["versoes_anteriores", "anteriores", "historico", "manual_cadip"]

ATTACH_CHANGED_FILES = True
MAX_ATTACHMENTS = 8
MAX_SINGLE_ATTACH_SIZE = 4 * 1024 * 1024
MAX_TOTAL_ATTACH_SIZE = 18 * 1024 * 1024

SEND_EMAIL_WHEN_NO_CHANGES = True  # fallback; tela: email.enviar_sem_alteracao


def _as_bool(valor, default: bool = False) -> bool:
    if isinstance(valor, bool):
        return valor
    if valor is None:
        return default
    return str(valor).strip().lower() in {"1", "true", "yes", "sim", "on"}


def _carregar_flags_monitor() -> tuple[bool, bool]:
    """Lê monitor.quiet_baseline / monitor.only_atual da tela (tabela configuracoes)."""
    quiet = QUIET_BASELINE
    only_atual = ONLY_ATUAL
    try:
        from persistencia.config_db import obter_configuracao

        quiet = _as_bool(obter_configuracao("monitor.quiet_baseline", quiet), quiet)
        only_atual = _as_bool(obter_configuracao("monitor.only_atual", only_atual), only_atual)
    except Exception as e:
        logger.warning(f"Config do monitor indisponível; usando defaults. Motivo: {e}")
    return quiet, only_atual


def _carregar_flags_email() -> tuple[bool, bool]:
    """Lê email.enviar_sem_alteracao / email.anexar_alterados da tela Configurações."""
    enviar_sem = SEND_EMAIL_WHEN_NO_CHANGES
    anexar = ATTACH_CHANGED_FILES
    try:
        from persistencia.config_db import obter_configuracao

        enviar_sem = _as_bool(
            obter_configuracao("email.enviar_sem_alteracao", enviar_sem), enviar_sem
        )
        anexar = _as_bool(obter_configuracao("email.anexar_alterados", anexar), anexar)
    except Exception as e:
        logger.warning(f"Config de e-mail (flags) indisponível; usando defaults. Motivo: {e}")
    return enviar_sem, anexar


def _metadados_mudaram(info: dict, anterior: dict, use_partial_fp: bool = True) -> tuple[bool, list[str]]:
    """Compara Last-Modified/etag/etc. com a baseline anterior. Sem anterior = sem mudança real."""
    if not anterior:
        return False, []
    reasons: list[str] = []
    for key in ("etag", "last_modified", "content_length", "final_url"):
        atual = info.get(key)
        if atual and atual != anterior.get(key):
            reasons.append(f"{key} mudou")
    if use_partial_fp:
        atual_fp = info.get("partial_fp")
        if atual_fp and atual_fp != anterior.get("partial_fp"):
            reasons.append("partial_fp mudou")
    return bool(reasons), reasons

# ====== PÁGINAS A MONITORAR ======
urls = [
    "https://www.bcb.gov.br/estabilidadefinanceira/leiautedocumentoDDR2011",
    "https://www.bcb.gov.br/estabilidadefinanceira/leiautedocumentoDRM",
    "https://www.bcb.gov.br/estabilidadefinanceira/leiautedoc2061",
    "https://www.bcb.gov.br/estabilidadefinanceira/leiautedoc2062",
    "https://www.bcb.gov.br/estabilidadefinanceira/leiaute_drl2160",
    "https://www.bcb.gov.br/estabilidadefinanceira/leiautedocumentoscrd",  # 4111 - SCD
]

# ====== DATA DE REFERÊNCIA ======
hoje = datetime.now().strftime("%d/%m/%Y")
ASSUNTO = f"📢 Atenção: Atualização na página de Leiautes do Bacen na data: {hoje}"

# ====== MANIFEST ======
MANIFEST_PATH = SCRIPT_DIR / "manifest_arquivos.json"

def _load_manifest():
    if MANIFEST_PATH.exists():
        try:
            return json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))
        except Exception as e:
            logger.warning(f"Falha ao ler manifest: {e}")
    return {}

def _save_manifest(data):
    MANIFEST_PATH.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")

# ====== REDE ======
def _session():
    sess = requests.Session()
    sess.headers.update({"User-Agent": "FINAUD-Monitor/1.0 (+https://local)"})
    return sess

def head_info(session, url):
    r = session.head(url, allow_redirects=True, timeout=TIMEOUT)
    r.raise_for_status()
    return {
        "etag": r.headers.get("ETag"),
        "last_modified": r.headers.get("Last-Modified"),
        "content_length": r.headers.get("Content-Length"),
        "final_url": r.url,
        "status": r.status_code,
        "checked_at": datetime.now().isoformat(),
    }

def small_range_fingerprint(session, url, length=1024):
    headers = {"Range": f"bytes=0-{length-1}"}
    r = session.get(url, headers=headers, stream=True, allow_redirects=True, timeout=TIMEOUT)
    if r.status_code not in (200, 206): r.raise_for_status()
    chunk = next(r.iter_content(length), b"")
    return hashlib.sha256(chunk).hexdigest()


def baixar_conteudo_para_historico(session, url, max_single=MAX_SINGLE_ATTACH_SIZE):
    try:
        hi = head_info(session, url)
        cl = hi.get("content_length")
        if cl and cl.isdigit() and int(cl) > max_single:
            return None, f"pula historico: Content-Length {cl} > limite"
    except Exception:
        pass

    r = session.get(url, stream=True, allow_redirects=True, timeout=TIMEOUT)
    if r.status_code != 200:
        return None, f"status {r.status_code}"

    data, total = bytearray(), 0
    for chunk in r.iter_content(64 * 1024):
        if not chunk:
            break
        data.extend(chunk)
        total += len(chunk)
        if total > max_single:
            return None, f"pula historico: excedeu {max_single} bytes"
    return bytes(data), None

# ====== ANEXOS ======
ANEXO_REGEX = re.compile(r"\.(pdf|xlsx?|xsd|zip)$", re.IGNORECASE)

def verificar_anexos(urls_anexos, categoria_por_url=None, execucao_id=None, use_partial_fp=True):
    """Compara anexos com a baseline anterior.

    Regra (item 3):
    - primeira observação de um URL: grava manifesto/baseline, sem evidência "Mudou";
    - evidência só quando Last-Modified/etag/hash mudou em relação à baseline.
    """
    quiet_baseline, _only_atual = _carregar_flags_monitor()
    manifest = _load_manifest()
    alterados, sess = [], _session()
    first_run = len(manifest) == 0
    categoria_por_url = categoria_por_url or {}
    primeiras_obs = 0

    if first_run:
        logger.info(
            "Manifesto vazio: primeira baseline. quiet_baseline=%s "
            "(evidências só em mudanças reais de metadados).",
            quiet_baseline,
        )

    for url in urls_anexos:
        cur = manifest.get(url, {})
        eh_primeira_obs = url not in manifest
        try:
            info = head_info(sess, url)
        except Exception as e:
            if use_partial_fp:
                try:
                    fp = small_range_fingerprint(sess, url)
                    info = {"etag":None,"last_modified":None,"content_length":None,
                            "final_url":url,"partial_fp":fp,"status":None,
                            "checked_at": datetime.now().isoformat()}
                except Exception as e2:
                    logger.warning(f"HEAD/Range falhou para {url}: {e2}")
                    manifest[url] = {**cur,"error": f"HEAD/Range fail: {e2}","checked_at": datetime.now().isoformat()}
                    continue
            else:
                logger.warning(f"HEAD falhou para {url}: {e}")
                manifest[url] = {**cur,"error": f"HEAD fail: {e}","checked_at": datetime.now().isoformat()}
                continue

        if not (info.get("etag") or info.get("last_modified") or info.get("content_length")) and use_partial_fp:
            if "partial_fp" not in info:
                try:
                    info["partial_fp"] = small_range_fingerprint(sess, url)
                except Exception:
                    pass

        mudanca_real, reasons = _metadados_mudaram(
            info, cur if not eh_primeira_obs else {}, use_partial_fp
        )
        evidencia = ""
        # Legado: arquivo novo na página (URL ainda não no manifesto) também alerta,
        # exceto na primeira baseline quieta do manifesto vazio.
        if eh_primeira_obs and not (first_run and quiet_baseline):
            mudanca_real = True
            if not reasons:
                reasons = ["novo arquivo observado"]
        if mudanca_real:
            logger.info(f"Alteração detectada em anexo: {url} | {', '.join(reasons)}")
            evidencia = ", ".join(reasons)
            alterados.append({"url": url, "evidencia": evidencia})
        elif eh_primeira_obs:
            primeiras_obs += 1
            logger.info(f"Baseline silenciosa (primeira observação): {url}")

        # Versão de conteúdo: mudança real, ou primeira obs (baseline sem evidência).
        precisa_versao = mudanca_real or eh_primeira_obs
        caminho_arquivo = None
        if salvar_conteudo_versao and precisa_versao:
            try:
                conteudo, motivo_historico = baixar_conteudo_para_historico(sess, url)
                if conteudo:
                    caminho_arquivo = salvar_conteudo_versao(
                        conteudo=conteudo,
                        nome_arquivo=_filename_from_url(url),
                        categoria=categoria_por_url.get(url),
                    )
                    logger.info(f"Versão salva para histórico: {caminho_arquivo}")
                elif motivo_historico:
                    logger.warning(f"Não foi possível salvar histórico de {url} | Motivo: {motivo_historico}")
            except Exception as e:
                logger.warning(f"Falha ao salvar versão no histórico: {url} | {e}")

        if registrar_arquivo_observado:
            try:
                registrar_arquivo_observado(
                    url=url,
                    nome_arquivo=_filename_from_url(url),
                    info=info,
                    categoria=categoria_por_url.get(url),
                    execucao_id=execucao_id,
                    mudou=precisa_versao,
                    evidencia=evidencia,
                    caminho_arquivo=caminho_arquivo,
                    gerar_evidencia=mudanca_real,
                )
            except Exception as e:
                logger.warning(f"Falha ao registrar arquivo no banco: {url} | {e}")

        manifest[url] = {
            "etag": info.get("etag"),
            "last_modified": info.get("last_modified"),
            "content_length": info.get("content_length"),
            "final_url": info.get("final_url"),
            "partial_fp": info.get("partial_fp") if use_partial_fp else cur.get("partial_fp"),
            "checked_at": info.get("checked_at"),
        }

    _save_manifest(manifest)
    if primeiras_obs:
        logger.info(
            "Primeiras observações sem evidência: %s arquivo(s). Mudanças reais: %s.",
            primeiras_obs,
            len(alterados),
        )
    return alterados, manifest, {
        "first_run": first_run,
        "quiet_baseline": quiet_baseline,
        "primeiras_observacoes": primeiras_obs,
    }


# ====== PLAYWRIGHT ======
def extrair_anexos_4111(page):
    linha_4111 = page.locator("tr").filter(has_text="4111")
    urls_4111, categorias = [], {}
    if linha_4111.count() > 0:
        links = linha_4111.locator("a")
        for i in range(links.count()):
            href = links.nth(i).get_attribute("href")
            if href:
                abs_url = page.evaluate("url => new URL(url, document.baseURI).toString()", href)
                if any(s in abs_url.lower() for s in [".pdf", ".xsd"]):
                    urls_4111.append(abs_url)
                    categorias[abs_url] = "4111 - SCD"
    return urls_4111, categorias


def _categoria_da_pagina(page_url: str) -> str:
    u = (page_url or "").lower()
    if "leiautedocumentoscrd" in u or "scrd" in u:
        return "4111 - SCD"
    if "ddr2011" in u or "documentoDDR".lower() in u:
        return "DDR-2011"
    if "leiautedocumentodrm" in u or "drm" in u:
        return "DRM-2060"
    if "leiautedoc2061" in u or "2061" in u:
        return "DLO-2061"
    if "leiautedoc2062" in u or "2062" in u:
        return "DLI-2062"
    if "drl2160" in u or "2160" in u:
        return "DRL-2160"
    return "Sem categoria"


def _filtrar_urls_anexos(candidatos: list[str]) -> list[str]:
    """Aplica excludes e only_atual.

    Páginas DLO/DLI usam pasta /Atual/. DRM/DRL não usam — nesses casos
    only_atual vira 'excluir histórico' (versoes_anteriores etc.), não exige /atual/.
    """
    limpos = []
    for u in candidatos:
        pl = (u or "").lower()
        if not pl:
            continue
        if any(pat in pl for pat in EXCLUDE_PATTERNS):
            continue
        limpos.append(u)

    if not ONLY_ATUAL:
        return limpos

    com_atual = [u for u in limpos if "/atual/" in u.lower()]
    if com_atual:
        return com_atual
    return limpos


def _coletar_hrefs_assets(page) -> list[str]:
    hrefs: list[str] = []
    try:
        hrefs.extend(
            page.evaluate(
                r"""() => {
                  const res = [];
                  const isAsset = (h) => /\.(pdf|xlsx?|xsd|zip)$/i.test(h||"");
                  for (const a of Array.from(document.querySelectorAll('a[href]'))) {
                    const href = a.getAttribute('href') || '';
                    if (!isAsset(href)) continue;
                    res.push(new URL(href, document.baseURI).toString());
                  }
                  return res;
                }"""
            )
        )
    except Exception:
        pass

    # Complemento: assets embutidos no HTML (ex.: XSD do DRM que às vezes não vira <a>).
    try:
        html = page.content()
        for m in re.findall(
            r"https?://[^\"'\s<>]+?\.(?:pdf|xlsx?|xsd|zip)",
            html,
            flags=re.I,
        ):
            hrefs.append(m)
        for m in re.findall(
            r"[\"']((?:/)?content/[^\"']+\.(?:pdf|xlsx?|xsd|zip))[\"']",
            html,
            flags=re.I,
        ):
            hrefs.append(page.evaluate("u => new URL(u, document.baseURI).toString()", m))
    except Exception:
        pass
    return list(dict.fromkeys(hrefs))


def extrair_datas_categorias_e_anexos(url):
    """Espelha o legado: espera conteúdo útil da página Angular e coleta anexos."""
    categoria_pagina = _categoria_da_pagina(url)
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=["--no-sandbox", "--disable-dev-shm-usage"])
        page = browser.new_page()
        # Páginas Bacen são Angular — networkidle + seletor de anexo evita lista vazia.
        page.goto(url, timeout=90000, wait_until="networkidle")
        try:
            page.wait_for_selector("table", timeout=8000)
        except Exception:
            pass
        try:
            page.wait_for_selector(
                'a[href$=".pdf"], a[href$=".xlsx"], a[href$=".xls"], '
                'a[href$=".xsd"], a[href$=".zip"], '
                'a[href*=".pdf?"], a[href*=".xlsx?"], a[href*=".xls?"]',
                timeout=25000,
            )
        except Exception:
            # Fallback: menu já tem <a>, mas conteúdo pode atrasar — dá tempo extra.
            page.wait_for_timeout(5000)

        if "leiautedocumentoscrd" in url.lower():
            anexos_4111, categorias_4111 = extrair_anexos_4111(page)
            browser.close()
            return [], anexos_4111, categorias_4111

        datas = []
        try:
            for cell in page.query_selector_all("td"):
                text = (cell.inner_text() or "").strip()
                if len(text) == 10 and text[2] == "/" and text[5] == "/":
                    datas.append(text)
        except Exception:
            pass

        hrefs = _coletar_hrefs_assets(page)
        # Retry curto se a hidratação ainda não trouxe anexos (caso 24/07 no 2061).
        if len(_filtrar_urls_anexos(hrefs)) < 3:
            page.wait_for_timeout(4000)
            hrefs = _coletar_hrefs_assets(page)

        browser.close()

        anexos = _filtrar_urls_anexos(hrefs)
        if not anexos:
            logger.warning(
                "Scrape sem anexos úteis em %s (possível página Angular incompleta).",
                url,
            )
        else:
            logger.info("Scrape %s: %s anexo(s) após filtro.", url, len(anexos))
        categoria_por_url = {u: categoria_pagina for u in anexos}
        return datas, anexos, categoria_por_url


# ====== EMAIL HTML ======
BLUE_BRAND = "#2e3192"


def _parse_json_lista(valor):
    if not valor:
        return []
    try:
        parsed = json.loads(valor)
        return parsed if isinstance(parsed, list) else []
    except Exception:
        return []


def _carregar_detalhes_alteracoes(execucao_id):
    if not execucao_id or not conectar or not init_db:
        return {}
    try:
        init_db()
        with conectar() as conn:
            rows = conn.execute(
                """
                SELECT
                    ar.url,
                    ar.nome_arquivo,
                    ar.tipo_arquivo,
                    COALESCE(l.codigo, '') AS leiaute_codigo,
                    a.resumo_executivo,
                    a.impacto_sugerido,
                    a.itens_incluidos,
                    a.itens_removidos,
                    a.itens_alterados
                FROM alteracoes_detectadas a
                JOIN arquivos_monitorados ar ON ar.id = a.arquivo_id
                LEFT JOIN leiautes_monitorados l ON l.id = ar.leiaute_id
                WHERE a.execucao_id = ?
                ORDER BY a.id DESC
                """,
                (execucao_id,),
            ).fetchall()
    except Exception as exc:
        logger.warning(f"Nao foi possivel carregar detalhes das alteracoes: {exc}")
        return {}

    detalhes = {}
    for row in rows:
        data = dict(row)
        data["itens_incluidos"] = _parse_json_lista(data.get("itens_incluidos"))
        data["itens_removidos"] = _parse_json_lista(data.get("itens_removidos"))
        data["itens_alterados"] = _parse_json_lista(data.get("itens_alterados"))
        detalhes[data["url"]] = data
    return detalhes


def _parse_evidencia_item(texto: str) -> dict:
    m = re.match(r"^Aba incluída: (.+)$", texto.strip())
    if m:
        nome = m.group(1).strip()
        return {"local": f"Aba {nome}", "depois": nome, "mudanca": "acrescentou aba"}
    m = re.match(r"^Aba removida: (.+)$", texto.strip())
    if m:
        nome = m.group(1).strip()
        return {"local": f"Aba {nome}", "antes": nome, "mudanca": "removeu aba"}
    m = re.match(r'^Aba renomeada: "(.+)" → "(.+)"$', texto.strip())
    if m:
        return {
            "local": f"Aba {m.group(1)}",
            "antes": m.group(1),
            "depois": m.group(2),
            "mudanca": "renomeou aba",
        }

    m = re.match(r'^Aba (.+), conta ([^:]+): incluída "(.+)"$', texto.strip())
    if m:
        return {
            "local": f"Aba {m.group(1).strip()}, conta {m.group(2).strip()}",
            "depois": m.group(3),
            "mudanca": "acrescentou conta",
        }
    m = re.match(r'^Aba (.+), conta ([^:]+): removida "(.+)"$', texto.strip())
    if m:
        return {
            "local": f"Aba {m.group(1).strip()}, conta {m.group(2).strip()}",
            "antes": m.group(3),
            "mudanca": "removeu conta",
        }

    padroes = [
        r'^(.*?): mudanca "([\s\S]*)"; antes "([\s\S]*)"; depois "([\s\S]*)"$',
        r'^(.*?): antes "([\s\S]*)"; depois "([\s\S]*)"$',
        r"^(.*?): antes '([\s\S]*)'; depois '([\s\S]*)'$",
        r"^(.*?): (linha anterior.*?); antes \((.*)\); depois \((.*)\)$",
        # XLSX/XLS do comparador: "Aba X, célula A1: antes valor; depois valor"
        r"^(.*?): antes ([\s\S]*); depois ([\s\S]*)$",
    ]
    for padrao in padroes:
        m = re.match(padrao, texto)
        if not m:
            continue
        if len(m.groups()) == 4 and "mudanca" in padrao:
            return {
                "local": m.group(1),
                "mudanca": m.group(2),
                "antes": m.group(3),
                "depois": m.group(4),
            }
        if len(m.groups()) == 4:
            return {"local": f"{m.group(1)} - {m.group(2)}", "antes": m.group(3), "depois": m.group(4)}
        return {"local": m.group(1), "antes": m.group(2), "depois": m.group(3)}

    m = re.match(r'^(.*?): incluído "([\s\S]*)"$', texto)
    if m:
        return {"local": m.group(1), "depois": m.group(2), "mudanca": f'acrescentou "{m.group(2)}"'}
    m = re.match(r'^(.*?): removido "([\s\S]*)"$', texto)
    if m:
        return {"local": m.group(1), "antes": m.group(2), "mudanca": f'removeu "{m.group(2)}"'}
    m = re.match(r'^Arquivo interno incluído: ([^;]+); evidência: "([\s\S]*)"$', texto)
    if m:
        return {"local": f"Arquivo interno {m.group(1)}", "depois": m.group(2)}
    return {"local": "Evidência", "depois": texto}


def _grupo_local(local: str) -> str:
    m = re.match(r"^(Página \d+)", local)
    if m:
        return m.group(1)
    m = re.match(r"^(Arquivo interno [^-:]+)", local)
    if m:
        return m.group(1).strip()
    if re.match(r"^linha (atual|anterior) \d+", local, flags=re.I):
        return "Linhas do arquivo"
    return local


def _numeros_linha(itens: list[dict], removido: bool = False) -> str:
    padrao = r"linha anterior (\d+)" if removido else r"linha atual (\d+)"
    numeros = []
    for item in itens:
        m = re.search(padrao, item.get("local", ""), flags=re.I)
        if m:
            numeros.append(int(m.group(1)))
    if not numeros:
        return ""
    numeros.sort()
    if len(numeros) == 1:
        return f"linha {numeros[0]}"
    return f"linhas {numeros[0]} a {numeros[-1]}"


def _inicia_paragrafo(linha: str) -> bool:
    return bool(
        re.match(r"^\d{2}/\d{2}/\d{4}", linha)
        or re.match(r"^Item\s+\d", linha, flags=re.I)
        or re.match(r"^[-•]", linha)
        or re.match(r"^Instrução Normativa", linha, flags=re.I)
        or re.match(r"^Resolução", linha, flags=re.I)
    )


def _limpar_texto_evidencia(texto: str) -> str:
    texto = re.sub(r"\s+", " ", str(texto)).strip()
    texto = re.sub(r"([^\W\d_])[-‐‑–]\s+([^\W\d_])", r"\1\2", texto, flags=re.UNICODE)
    texto = re.sub(r"\s+([,.;:])", r"\1", texto)
    return texto


def _consolidar_textos(textos: list[str]) -> tuple[list[str], int]:
    paragrafos = []
    atual = ""
    extras = 0

    def flush():
        nonlocal atual
        if atual.strip():
            paragrafos.append(_limpar_texto_evidencia(atual))
        atual = ""

    for bruto in textos:
        linha = _limpar_texto_evidencia(str(bruto))
        if not linha:
            continue
        mais = re.match(r"^\.\.\. mais (\d+)", linha, flags=re.I)
        if mais:
            extras += int(mais.group(1))
            continue
        if _inicia_paragrafo(linha):
            flush()
            atual = linha
        elif atual and (not re.search(r"[.!?:;)]$", atual) or re.match(r"^[a-záéíóúàâêôãõç]", linha)):
            atual = f"{atual} {linha}"
        else:
            flush()
            atual = linha
    flush()
    return paragrafos, extras


def _html_card_simples(titulo: str, textos: list[str], removido: bool = False) -> str:
    paragrafos, extras = _consolidar_textos(textos)
    corpo = "".join(f"<p>{html.escape(p)}</p>" for p in paragrafos)
    if extras:
        corpo += f"<p class='more'>+ {extras} trecho(s) adicional(is) no arquivo.</p>"
    return f"""
      <div class="evidence-card">
        <div class="evidence-head"><strong>{html.escape(titulo)}</strong></div>
        <div class="evidence-body">{corpo}</div>
      </div>
    """


# Limite de linhas de evidência por arquivo no e-mail (leitura limpa).
# A lista completa vai na planilha anexa Antes_Depois_leiautes_*.xlsx.
MAX_DIFFS_EMAIL = 5
ANEXO_ANTES_DEPOIS_PREFIXO = "Antes_Depois_leiautes"

def _eh_novo_arquivo(texto: str) -> bool:
    t = (texto or "").lower()
    return "novo arquivo observado" in t or "arquivo novo" in t


def _eh_evidencia_tecnica(texto: str) -> bool:
    """Metadados HTTP / republicação — não inclui arquivo novo na página."""
    if _eh_novo_arquivo(texto):
        return False
    t = (texto or "").lower()
    chaves = (
        "etag",
        "last_modified",
        "content_length",
        "final_url",
        "partial_fp",
        "metadados",
        "versão anterior não arquivada",
        "versao anterior nao arquivada",
        "alteracao detectada por metadados",
        "alteração detectada por metadados",
    )
    return any(k in t for k in chaves)


def _separar_itens_tecnicos_e_conteudo(itens: list[str]) -> tuple[list[str], list[str]]:
    tecnicos, conteudo = [], []
    for bruto in itens or []:
        texto = str(bruto)
        if _eh_evidencia_tecnica(texto):
            tecnicos.append(texto)
        else:
            conteudo.append(texto)
    return tecnicos, conteudo


def _detalhe_so_tecnico(detalhe: dict | None, evidencia: str = "") -> bool:
    # Arquivo novo na página do Bacen = gestor precisa revisar (como o legado alerta).
    if _eh_novo_arquivo(evidencia):
        return False
    if detalhe:
        resumo = str(detalhe.get("resumo_executivo") or "")
        if _eh_novo_arquivo(resumo):
            return False
        for lista in (
            detalhe.get("itens_incluidos") or [],
            detalhe.get("itens_removidos") or [],
            detalhe.get("itens_alterados") or [],
        ):
            if any(_eh_novo_arquivo(str(x)) for x in lista):
                return False

    if not detalhe:
        return bool(evidencia) and _eh_evidencia_tecnica(evidencia)
    inc = detalhe.get("itens_incluidos") or []
    rem = detalhe.get("itens_removidos") or []
    alt = detalhe.get("itens_alterados") or []
    _, conteudo_alt = _separar_itens_tecnicos_e_conteudo(alt)
    if inc or rem or conteudo_alt:
        return False
    resumo = str(detalhe.get("resumo_executivo") or "")
    if _eh_evidencia_tecnica(resumo) or _eh_evidencia_tecnica(evidencia):
        return True
    # Comparador sem diff de conteúdo, só sinal de mudança no site.
    if "nenhuma diferença" in resumo.lower() or "nenhuma diferenca" in resumo.lower():
        return True
    return not (inc or rem or alt)


def _contagem_curta(detalhe: dict | None, evidencia: str = "") -> str:
    if _eh_novo_arquivo(evidencia) or (
        detalhe
        and (
            _eh_novo_arquivo(str(detalhe.get("resumo_executivo") or ""))
            or any(
                _eh_novo_arquivo(str(x))
                for x in (detalhe.get("itens_incluidos") or [])
                + (detalhe.get("itens_alterados") or [])
            )
        )
    ):
        return "arquivo novo"
    if _detalhe_so_tecnico(detalhe, evidencia):
        return "técnico"
    if not detalhe:
        return evidencia or "alterado"
    _, conteudo_alt = _separar_itens_tecnicos_e_conteudo(detalhe.get("itens_alterados") or [])
    n_in = len(detalhe.get("itens_incluidos") or [])
    n_out = len(detalhe.get("itens_removidos") or [])
    n_ch = len(conteudo_alt)
    partes = []
    if n_ch:
        partes.append(f"{n_ch} mudou")
    if n_in:
        partes.append(f"{n_in} entrou")
    if n_out:
        partes.append(f"{n_out} saiu")
    if not partes:
        return "conteúdo alterado"
    return " · ".join(partes)


def _html_marcar_diferenca(antes: str, depois: str) -> tuple[str, str]:
    """Destaca em vermelho/verde só o trecho que mudou (e-mail)."""
    a = antes or ""
    b = depois or ""
    if a == b:
        return html.escape(a or "—"), html.escape(b or "—")
    sm = difflib.SequenceMatcher(None, a, b, autojunk=False)
    ha: list[str] = []
    hb: list[str] = []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        ta, tb = html.escape(a[i1:i2]), html.escape(b[j1:j2])
        if tag == "equal":
            ha.append(ta)
            hb.append(tb)
        elif tag == "replace":
            if ta:
                ha.append(f'<span style="background:#ffe0e0;color:#9b0000;">{ta}</span>')
            if tb:
                hb.append(f'<span style="background:#e3f5e3;color:#0b6b0b;">{tb}</span>')
        elif tag == "delete":
            if ta:
                ha.append(f'<span style="background:#ffe0e0;color:#9b0000;">{ta}</span>')
        elif tag == "insert":
            if tb:
                hb.append(f'<span style="background:#e3f5e3;color:#0b6b0b;">{tb}</span>')
    return "".join(ha) or "—", "".join(hb) or "—"


def _html_tabela_mudancas(itens: list[str]) -> str:
    """Tabela Onde | O que mudou | Antes | Depois — a coluna do meio é a que o gestor lê."""
    if not itens:
        return ""
    linhas = []
    for texto in itens[:MAX_DIFFS_EMAIL]:
        item = _parse_evidencia_item(str(texto))
        local = html.escape(str(item.get("local") or "Alteração"))
        antes_bruto = str(item.get("antes") or "—")
        depois_bruto = str(item.get("depois") or "—")
        mudanca = str(item.get("mudanca") or "").strip()
        if not mudanca and (antes_bruto != "—" or depois_bruto != "—"):
            try:
                from backend.app.services.comparador_arquivos import _descrever_mudanca

                mudanca = _descrever_mudanca(
                    "" if antes_bruto == "—" else antes_bruto,
                    "" if depois_bruto == "—" else depois_bruto,
                )
            except Exception:
                mudanca = "texto alterado"
        if len(antes_bruto) > 160 or len(depois_bruto) > 160:
            try:
                from backend.app.services.comparador_arquivos import _recortar_par_diff

                antes_bruto, depois_bruto = _recortar_par_diff(antes_bruto, depois_bruto)
            except Exception:
                antes_bruto, depois_bruto = antes_bruto[:160] + "…", depois_bruto[:160] + "…"
        antes_h, depois_h = _html_marcar_diferenca(antes_bruto, depois_bruto)
        mudanca_h = (
            f'<strong style="color:#2e3192;">{html.escape(mudanca)}</strong>'
            if mudanca
            else "—"
        )
        linhas.append(
            f"<tr><td class='col-local'>{local}</td>"
            f"<td>{mudanca_h}</td>"
            f"<td>{antes_h}</td><td>{depois_h}</td></tr>"
        )
    extra = ""
    if len(itens) > MAX_DIFFS_EMAIL:
        extra = (
            f"<p class='more'>+ {len(itens) - MAX_DIFFS_EMAIL} "
            f"alteração(ões) adicional(is) — lista completa na planilha anexa "
            f"<strong>{html.escape(ANEXO_ANTES_DEPOIS_PREFIXO)}_….xlsx</strong> "
            f"(colunas Onde / O que mudou / Antes / Depois / O que fazer).</p>"
        )
    return f"""
      <table class="diff-table" role="presentation" cellpadding="0" cellspacing="0">
        <thead>
          <tr>
            <th>Onde</th><th>O que mudou</th><th>Antes</th><th>Depois</th>
          </tr>
        </thead>
        <tbody>
          {''.join(linhas)}
        </tbody>
      </table>
      {extra}
    """


def _html_lista_simples(titulo: str, tipo: str, itens: list[str]) -> str:
    """Lista curta para inclusões/remoções — omitida se vazia."""
    if not itens:
        return ""
    bullets = []
    for texto in itens[:MAX_DIFFS_EMAIL]:
        item = _parse_evidencia_item(str(texto))
        if tipo == "saiu":
            trecho = item.get("antes") or item.get("depois") or texto
        else:
            trecho = item.get("depois") or item.get("antes") or texto
        local = str(item.get("local") or "Item")
        if local.startswith("Aba ") and trecho == local[4:]:
            bullets.append(f"<li><strong>{html.escape(trecho)}</strong></li>")
        else:
            bullets.append(
                f"<li><strong>{html.escape(local)}:</strong> "
                f"{html.escape(str(trecho))}</li>"
            )
    extra = ""
    if len(itens) > MAX_DIFFS_EMAIL:
        extra = (
            f"<p class='more'>+ {len(itens) - MAX_DIFFS_EMAIL} item(ns) no e-mail; "
            f"lista completa na planilha anexa "
            f"<strong>{html.escape(ANEXO_ANTES_DEPOIS_PREFIXO)}_….xlsx</strong>.</p>"
        )
    return f"""
      <p class="sec-label">{html.escape(titulo)}</p>
      <ul class="compact-list">{''.join(bullets)}</ul>
      {extra}
    """


def _html_lista_diferencas(titulo: str, tipo: str, itens: list[str], vazio: str = "") -> str:
    """Compat: seções vazias não aparecem (e-mail limpo)."""
    del vazio  # não mostrar estado vazio
    if not itens:
        return ""
    if tipo == "mudou":
        return f'<p class="sec-label">{html.escape(titulo)}</p>{_html_tabela_mudancas(itens)}'
    return _html_lista_simples(titulo, tipo, itens)


def _tem_diff_conteudo(detalhe: dict | None) -> bool:
    """True se já há Antes/Depois (células/texto) para o gestor ler no e-mail."""
    if not detalhe:
        return False
    inc = detalhe.get("itens_incluidos") or []
    rem = detalhe.get("itens_removidos") or []
    _, conteudo = _separar_itens_tecnicos_e_conteudo(
        detalhe.get("itens_alterados") or []
    )
    conteudo = [
        c for c in conteudo if "novo arquivo observado" not in str(c).lower()
    ]
    inc = [i for i in inc if "novo arquivo na página" not in str(i).lower()]
    return bool(inc or rem or conteudo)


def _tipo_comparacao_rotulo(detalhe: dict | None) -> tuple[str, str]:
    """Retorna (código, rótulo amigável) a partir do resumo gravado."""
    resumo = str((detalhe or {}).get("resumo_executivo") or "")
    if resumo.startswith("[Versão pareada"):
        # Extrai "antes: Nome" se houver
        m = re.search(r"antes:\s*([^\]]+)\]", resumo)
        antes = (m.group(1).strip() if m else "").strip()
        if antes:
            return (
                "versao_pareada",
                f"Nova versão na página (arquivo novo) — Antes: {antes}",
            )
        return "versao_pareada", "Nova versão na página (arquivo novo)"
    if resumo.startswith("[Sem anterior]"):
        return "sem_anterior", "Arquivo novo — sem versão anterior para comparar"
    if resumo.startswith("[Mesmo arquivo]"):
        return "mesmo_arquivo", "Atualização do mesmo arquivo"
    return "desconhecido", "Comparação automática"


def _pattern_so_digitos(texto: str) -> bool:
    t = texto or ""
    if re.search(r"\[0-9\]\{8\}", t) or re.search(r"\[0-9\]\[0-9\]\[0-9\]\[0-9\]", t):
        return True
    if re.search(r"\[a-zA-Z0-9\]", t, flags=re.I):
        return False
    return bool(re.search(r"pattern[^>]*\[0-9\]", t, flags=re.I))


def _pattern_alfanumerico(texto: str) -> bool:
    return bool(re.search(r"\[a-zA-Z0-9\]", texto or "", flags=re.I))


def _extrair_codigos_pipe(texto: str) -> set[str]:
    return {c for c in re.findall(r"\|([A-Za-z0-9_]{1,12})\|", texto or "")}


def _explicar_item_simples(texto: str, tipo_arquivo: str) -> tuple[str, int] | None:
    """Traduz um item de diff → (frase, prioridade). Prioridade menor = mais importante."""
    item = _parse_evidencia_item(str(texto))
    antes = str(item.get("antes") or "").strip()
    depois = str(item.get("depois") or "").strip()
    local = str(item.get("local") or "").strip()
    mudanca = str(item.get("mudanca") or "").strip().lower()
    junto = f"{local} {antes} {depois} {mudanca}".lower()

    # Metadado de pareamento (já coberto pelo tipo de comparação / cabeçalho).
    if "versão pareada" in mudanca or "versao pareada" in mudanca:
        return None

    # Vigência / data-base (comentários XSD ou texto).
    if "data-base" in junto or "data base" in junto or "competência" in junto:
        m_ant = re.search(
            r"(\d{2}/\d{4}|\d{2}/\d{2}/\d{4}|20\d{2}\d{2})", antes
        )
        m_dep = re.search(
            r"(\d{2}/\d{4}|\d{2}/\d{2}/\d{4}|20\d{2}\d{2})", depois
        )
        if m_ant and m_dep and m_ant.group(1) != m_dep.group(1):
            return (
                f"A vigência/data-base muda de {m_ant.group(1)} para {m_dep.group(1)}.",
                1,
            )
        if m_dep:
            return (f"A vigência/data-base passa a {m_dep.group(1)}.", 1)

    # Só etiqueta de versão / “Atualizado em”.
    if (
        re.search(r"vers[aã]o\s*[\d.]+", antes, flags=re.I)
        or "<!-- + versao" in antes.lower()
        or "<!-- + versão" in antes.lower()
    ) and (
        "versao" in depois.lower() or "versão" in depois.lower()
    ):
        return ("Atualizaram o rótulo interno de versão (só etiqueta).", 8)
    if "atualizado em" in junto and re.search(r"\d{2}/\d{2}/\d{4}", junto):
        return ("Atualizaram a data de publicação no cabeçalho (só etiqueta).", 8)

    # Pattern XSD: numérico → alfanumérico (e vice-versa).
    if "pattern" in junto or _pattern_so_digitos(antes) or _pattern_alfanumerico(depois):
        if _pattern_so_digitos(antes) and _pattern_alfanumerico(depois):
            return (
                "Um campo que só aceitava números agora aceita letras e números "
                "(código alfanumérico).",
                0,
            )
        if _pattern_alfanumerico(antes) and _pattern_so_digitos(depois):
            return (
                "Um campo que aceitava letras/números passou a aceitar só números.",
                0,
            )
        # Códigos novos dentro do pattern (ex.: acrescentou "|B1").
        fonte = f"{mudanca}\n{texto}"
        m_acresc = re.findall(
            r'acrescentou\s+"\|?([A-Za-z0-9_]{1,12})\|?"',
            fonte,
            flags=re.I,
        )
        unicos: list[str] = []
        vistos_ci: set[str] = set()
        for c in m_acresc:
            if c.upper() in {"S", "N"} or c.isdigit() or len(c) > 6:
                continue
            if c.upper() in vistos_ci:
                continue
            vistos_ci.add(c.upper())
            unicos.append(c.upper())
        if unicos and len(unicos) <= 6:
            return (
                f"Passa a aceitar o(s) código(s): {', '.join(unicos)}.",
                0,
            )
        if "pattern" in junto and antes and depois and antes != depois:
            return (
                "Mudou a regra de formato de um campo (validação do envio).",
                1,
            )

    # Novos códigos em enumeração (ex.: |B1|).
    cod_ant = _extrair_codigos_pipe(antes)
    cod_dep = _extrair_codigos_pipe(depois)
    if not (cod_ant or cod_dep):
        m_ad = re.search(
            r'antes\s+"([\s\S]*?)"\s*;\s*depois\s+"([\s\S]*)"\s*$',
            str(texto),
            flags=re.I,
        )
        if m_ad:
            cod_ant = _extrair_codigos_pipe(m_ad.group(1))
            cod_dep = _extrair_codigos_pipe(m_ad.group(2))
    novos = sorted(cod_dep - cod_ant)
    removidos = sorted(cod_ant - cod_dep)
    if novos and len(novos) <= 8:
        lista = ", ".join(novos[:6])
        extra = f" (+{len(novos) - 6})" if len(novos) > 6 else ""
        return (f"Passa a aceitar o(s) código(s): {lista}{extra}.", 0)
    if removidos and len(removidos) <= 8:
        lista = ", ".join(removidos[:6])
        return (f"Deixa de aceitar o(s) código(s): {lista}.", 0)
    m_acresc = re.findall(
        r'acrescentou\s+"\|?([A-Za-z0-9_]{1,12})\|?"',
        mudanca or str(texto),
        flags=re.I,
    )
    if m_acresc and ("enumeration" in junto or "xs:" in junto or tipo_arquivo == "xsd"):
        unicos = sorted(set(m_acresc))
        if len(unicos) <= 6:
            return (f"Passa a aceitar o(s) código(s): {', '.join(unicos)}.", 0)

    # Planilha: célula / aba.
    if tipo_arquivo in {"xlsx", "xls", "xlsm"} or "célula" in local.lower() or "aba " in local.lower():
        if "acrescentou aba" in mudanca or mudanca == "acrescentou aba":
            return (f'Entrou a aba "{depois or local}".', 1)
        if "removeu aba" in mudanca:
            return (f'Saiu a aba "{antes or local}".', 1)
        if "renomeou aba" in mudanca:
            return (f'A aba "{antes}" passou a se chamar "{depois}".', 1)
        if "conta" in mudanca:
            if "acrescentou" in mudanca:
                return (f'Entrou a conta "{depois}" ({local}).', 1)
            if "removeu" in mudanca:
                return (f'Saiu a conta "{antes}" ({local}).', 1)
        if antes and depois and antes != depois:
            local_curto = local if len(local) <= 60 else local[:57] + "…"
            a = antes if len(antes) <= 40 else antes[:37] + "…"
            d = depois if len(depois) <= 40 else depois[:37] + "…"
            return (f"Na planilha ({local_curto}): \"{a}\" -> \"{d}\".", 1)

    # PDF: página.
    if tipo_arquivo == "pdf" or local.lower().startswith("página"):
        m_pag = re.search(r"página\s+(\d+)", local, flags=re.I)
        pag = f" na página {m_pag.group(1)}" if m_pag else ""
        if antes and depois:
            return (f"Mudou o texto das instruções{pag}.", 2)
        if depois and not antes:
            return (f"Entrou texto novo nas instruções{pag}.", 2)
        if antes and not depois:
            return (f"Saiu texto das instruções{pag}.", 2)

    # Inclusão/remoção genérica de campo/schema.
    if "incluíd" in junto or "incluído" in junto or "incluido" in junto:
        trecho = depois or local
        trecho = trecho if len(trecho) <= 70 else trecho[:67] + "…"
        return (f"Entrou algo novo no arquivo: {trecho}.", 3)
    if "removid" in junto:
        trecho = antes or local
        trecho = trecho if len(trecho) <= 70 else trecho[:67] + "…"
        return (f"Saiu algo do arquivo: {trecho}.", 3)

    return None


def _explicar_linguagem_simples(
    detalhe: dict | None, item: dict | None = None
) -> list[str]:
    """Gera bullets em português claro a partir do diff estruturado (só regras)."""
    det = detalhe or {}
    nome = str(det.get("nome_arquivo") or "")
    if not nome and item:
        nome = _filename_from_url(item.get("url") or "")
    tipo = str(det.get("tipo_arquivo") or Path(nome).suffix.lstrip(".")).lower()
    tipo_cmp, _rotulo = _tipo_comparacao_rotulo(det)

    frases: list[tuple[int, str]] = []
    vistos: set[str] = set()

    def _add(frase: str | None, prio: int) -> None:
        if not frase:
            return
        chave = frase.strip().lower()
        if chave in vistos:
            return
        vistos.add(chave)
        frases.append((prio, frase.strip()))

    if tipo_cmp == "versao_pareada":
        m = re.search(r"antes:\s*([^\]]+)\]", str(det.get("resumo_executivo") or ""))
        antes_nome = (m.group(1).strip() if m else "").strip()
        if antes_nome:
            _add(
                f"Arquivo novo comparado com a versão anterior ({antes_nome}).",
                2,
            )
        else:
            _add("Arquivo novo comparado com a versão anterior do mesmo documento.", 2)
    elif tipo_cmp == "sem_anterior":
        _add("Arquivo novo sem versão anterior para comparar automaticamente.", 2)
    elif tipo_cmp == "mesmo_arquivo":
        _add("O mesmo arquivo foi atualizado no site do Bacen.", 4)

    todos = list(det.get("itens_incluidos") or []) + list(
        det.get("itens_removidos") or []
    ) + list(det.get("itens_alterados") or [])
    for bruto in todos:
        if "novo arquivo na página" in str(bruto).lower():
            continue
        if "novo arquivo observado" in str(bruto).lower():
            continue
        explicado = _explicar_item_simples(str(bruto), tipo)
        if explicado:
            frase, prio = explicado
            _add(frase, prio)

    if not any(p <= 3 for p, _ in frases) and _tem_diff_conteudo(det):
        _add(
            "Há diferenças de conteúdo; confira a tabela Antes/Depois abaixo.",
            5,
        )

    frases.sort(key=lambda x: (x[0], x[1]))
    # Se já há impacto real, omite só-etiqueta.
    if any(p <= 1 for p, _ in frases):
        frases = [(p, f) for p, f in frases if p < 8]
    # Máx. 6 bullets; prioriza impacto.
    saida = [f for _, f in frases[:6]]
    if not saida and _detalhe_so_tecnico(det, (item or {}).get("evidencia") or ""):
        saida = [
            "O Bacen republicou o arquivo no site — em geral sem ação sua."
        ]
    return saida


def _html_linguagem_simples(bullets: list[str]) -> str:
    if not bullets:
        return ""
    lis = "".join(f"<li>{html.escape(b)}</li>" for b in bullets)
    return f"""
      <div class="simples">
        <p class="simples-title">Em linguagem simples</p>
        <ul class="simples-list">{lis}</ul>
      </div>
    """


def _texto_situacao(item: dict, detalhe: dict | None) -> str:
    """Passo 1: o que aconteceu, sem jargão."""
    det = detalhe or {}
    codigo = str(det.get("leiaute_codigo") or "").strip()
    nome = str(det.get("nome_arquivo") or _filename_from_url(item.get("url") or ""))
    tipo_cmp, _ = _tipo_comparacao_rotulo(det)
    leiaute = f" do leiaute {codigo}" if codigo else ""
    m = re.search(r"antes:\s*([^\]]+)\]", str(det.get("resumo_executivo") or ""))
    antes_nome = (m.group(1).strip() if m else "").strip()

    if tipo_cmp == "versao_pareada":
        if antes_nome:
            return (
                f"Apareceu um arquivo novo{leiaute} na página do Bacen: {nome}. "
                f"Para você ver o que mudou, o robô compara com a versão anterior "
                f"do mesmo documento ({antes_nome}). "
                "Não é o mesmo arquivo republicado: é a versão nova ao lado da antiga."
            )
        return (
            f"Apareceu um arquivo novo{leiaute} na página do Bacen: {nome}. "
            "Para você ver o que mudou, o robô compara com a versão anterior "
            "do mesmo documento."
        )
    if tipo_cmp == "sem_anterior":
        return (
            f"Apareceu um arquivo novo{leiaute} na página do Bacen: {nome}. "
            "Ainda não há versão anterior no histórico para comparar automaticamente."
        )
    return (
        f"O arquivo{leiaute} já monitorado foi atualizado no site do Bacen: {nome}."
    )


def _texto_o_que_fazer(item: dict, detalhe: dict | None) -> str:
    """Passo 3: ação clara, sem jargão."""
    det = detalhe or {}
    impacto = str(det.get("impacto_sugerido") or "").strip()
    evidencia = item.get("evidencia") or ""
    resumo = str(det.get("resumo_executivo") or "")
    codigo = str(det.get("leiaute_codigo") or "").strip()
    nome = str(det.get("nome_arquivo") or _filename_from_url(item.get("url") or ""))
    tipo = str(det.get("tipo_arquivo") or Path(nome).suffix.lstrip(".")).lower()
    alvo = f" do leiaute {codigo}" if codigo else ""
    tem_diff = _tem_diff_conteudo(det)
    tipo_cmp, _rotulo = _tipo_comparacao_rotulo(det)
    bullets = _explicar_linguagem_simples(det, item)
    # Frases de impacto real (ignora cabeçalho de arquivo novo).
    foco = [
        b
        for b in bullets
        if not b.lower().startswith("arquivo novo")
        and "versão anterior" not in b.lower()
        and "versao anterior" not in b.lower()
        and "atualização do mesmo arquivo" not in b.lower()
        and "atualizacao do mesmo arquivo" not in b.lower()
    ]

    if tem_diff:
        if foco:
            resumo_foco = " ".join(foco[:3])
            return (
                f"Verifique se a rotina{alvo} já cobre o que mudou: {resumo_foco} "
                "Os detalhes completos estão na planilha em anexo."
            )
        if tipo in {"xlsx", "xls", "xlsm"}:
            return (
                f"Atualize parâmetros e limites da rotina interna{alvo} "
                "conforme a coluna Depois da tabela. Detalhes na planilha em anexo."
            )
        if tipo == "pdf":
            return (
                f"Ajuste o preenchimento/rotina{alvo} conforme o texto Depois "
                "nas diferenças acima. Detalhes na planilha em anexo."
            )
        if tipo == "xsd":
            return (
                f"Confira se os sistemas de envio{alvo} precisam ser atualizados "
                "com as mudanças acima. Detalhes na planilha em anexo."
            )
        return (
            impacto
            if impacto
            and "compare" not in impacto.lower()
            and "url" not in impacto.lower()
            else (
                f"Revise as diferenças acima{alvo} e atualize a rotina interna "
                "se necessário. Detalhes na planilha em anexo."
            )
        )

    if tipo_cmp == "sem_anterior" or (
        (_eh_novo_arquivo(evidencia) or _eh_novo_arquivo(resumo))
        and tipo_cmp != "versao_pareada"
    ):
        if tipo_cmp == "sem_anterior" or "[Sem anterior]" in resumo:
            return (
                f"Arquivo novo{alvo} sem versão anterior para comparar. "
                "Ainda não há tabela Antes/Depois automática."
            )
        return (
            f"Arquivo novo{alvo} detectado, mas a comparação automática "
            "ainda não ficou pronta neste alerta."
        )
    if tipo_cmp == "versao_pareada":
        return (
            impacto
            if impacto and "url" not in impacto.lower()
            else (
                f"Arquivo novo{alvo} comparado com a versão anterior; "
                "não houve mudança relevante de conteúdo."
            )
        )
    if impacto and "compare" not in impacto.lower() and "url" not in impacto.lower():
        return impacto
    return (
        f"Mudança{alvo} sem tabela Antes/Depois automática neste alerta."
    )


def _html_passo(numero: int, titulo: str, corpo: str) -> str:
    return f"""
      <div class="passo">
        <p class="passo-titulo">Passo {numero} de 3 · {html.escape(titulo)}</p>
        {corpo}
      </div>
    """


def _html_detalhe_conteudo(item, detalhe) -> str:
    """Arquivo em 3 passos: situação → diferenças → ação."""
    url = item["url"]
    nome = _filename_from_url(url)
    evidencia = item.get("evidencia") or ""
    codigo = ((detalhe or {}).get("leiaute_codigo") or "").strip()
    titulo_nome = (detalhe or {}).get("nome_arquivo") or nome
    rotulo = f"{codigo} · {titulo_nome}" if codigo else str(titulo_nome)
    contagem = _contagem_curta(detalhe, evidencia)
    link = (
        f'<a href="{html.escape(url)}" target="_blank">{html.escape(rotulo)}</a>'
    )

    situacao = _texto_situacao(item, detalhe)
    o_que_fazer = _texto_o_que_fazer(item, detalhe)
    simples_html = _html_linguagem_simples(
        _explicar_linguagem_simples(detalhe, item)
    )

    incluidos = [
        i
        for i in ((detalhe or {}).get("itens_incluidos") or [])
        if "novo arquivo na página" not in str(i).lower()
    ]
    removidos = (detalhe or {}).get("itens_removidos") or []
    _, conteudo = _separar_itens_tecnicos_e_conteudo(
        (detalhe or {}).get("itens_alterados") or []
    )
    conteudo = [
        c
        for c in conteudo
        if "novo arquivo observado" not in str(c).lower()
    ]
    # Linha "Comparação: versão pareada" é técnica; a situação já explica.
    conteudo = [
        c
        for c in conteudo
        if "versão pareada" not in str(c).lower()
        and "versao pareada" not in str(c).lower()
    ]
    secoes = "".join(
        [
            _html_lista_diferencas("Entrou", "entrou", incluidos),
            _html_lista_diferencas("Saiu", "saiu", removidos),
            _html_lista_diferencas("Mudou", "mudou", conteudo),
        ]
    )
    resumo = str((detalhe or {}).get("resumo_executivo") or "").strip()
    if not secoes and (_eh_novo_arquivo(evidencia) or _eh_novo_arquivo(resumo)):
        secoes = (
            "<p class='desc' style='margin:0'>"
            "Ainda não há tabela Antes/Depois automática neste item."
            "</p>"
        )

    passo1 = _html_passo(
        1,
        "Situação",
        f"<p class='desc' style='margin:0'>{html.escape(situacao)}</p>"
        "<p class='passo-next'>A seguir: o que mudou entre as versões.</p>",
    )
    meio = secoes + simples_html
    if not meio.strip():
        meio = "<p class='desc' style='margin:0'>Sem diferenças listadas.</p>"
    passo2 = _html_passo(
        2,
        "Diferenças",
        meio
        + "<p class='passo-next'>A seguir: o que fazer com isso.</p>",
    )
    passo3 = _html_passo(
        3,
        "Ação",
        f'<p class="acao" style="margin:0"><strong>O que fazer:</strong> '
        f"{html.escape(o_que_fazer)}</p>",
    )

    return f"""
      <div class="file-block">
        <p class="file-title">{link}
          <span class="muted"> — {html.escape(contagem)}</span></p>
        {passo1}
        {passo2}
        {passo3}
      </div>
    """


def _descrever_mudanca_tecnica(evidencia: str = "", detalhe: dict | None = None) -> str:
    """Traduz evidência técnica (etag/lm/…) para frase que o gestor entende."""
    textos = [evidencia or ""]
    if detalhe:
        textos.append(str(detalhe.get("resumo_executivo") or ""))
        for lista in (
            detalhe.get("itens_alterados") or [],
            detalhe.get("itens_incluidos") or [],
        ):
            textos.extend(str(x) for x in lista)
    junto = " ".join(textos).lower()

    partes: list[str] = []
    if "last_modified" in junto or "last-modified" in junto:
        partes.append("data de publicação no site")
    if "content_length" in junto:
        partes.append("tamanho do arquivo no site")
    if "etag" in junto:
        partes.append("identificador interno do arquivo no site")
    if "final_url" in junto:
        partes.append("endereço final do download")
    if "partial_fp" in junto:
        partes.append("assinatura parcial do arquivo")
    if "versão anterior não arquivada" in junto or "versao anterior nao arquivada" in junto:
        partes.append("sem versão anterior arquivada para comparar")
    if not partes and ("metadado" in junto or "republic" in junto):
        partes.append("republicação no site (metadados)")
    if not partes:
        # Fallback: limpa o jargão cru.
        bruto = (evidencia or "").strip()
        if bruto:
            return bruto.replace(" mudou", "").replace("_", " ").strip() or "metadados no site"
        return "metadados no site"
    if len(partes) == 1:
        return f"mudou a {partes[0]}" if not partes[0].startswith("sem ") else partes[0]
    return "mudou: " + "; ".join(partes)


def _html_lista_tecnica(itens_bloco: list[tuple[dict, dict]]) -> str:
    """Lista técnica com o que mudou no site (ainda sem ação de conteúdo)."""
    bullets = []
    for item, detalhe in itens_bloco:
        url = item["url"]
        nome = _filename_from_url(url)
        codigo = (detalhe.get("leiaute_codigo") or "").strip()
        rotulo = f"{codigo} · {nome}" if codigo else nome
        link = (
            f'<a href="{html.escape(url)}" target="_blank">'
            f"{html.escape(rotulo)}</a>"
        )
        o_que = _descrever_mudanca_tecnica(item.get("evidencia") or "", detalhe)
        bullets.append(
            f"<li>{link}"
            f"<br><span class='muted'><strong>O que mudou:</strong> "
            f"{html.escape(o_que)}</span></li>"
        )
    return f"<ul class='tech-list'>{''.join(bullets)}</ul>"


def _html_detalhe_alteracao(item, detalhe) -> str:
    """Compat: detalhe de conteúdo; técnico vira item de lista no bloco."""
    evidencia = item.get("evidencia") or ""
    if _detalhe_so_tecnico(detalhe, evidencia):
        return _html_lista_tecnica([(item, detalhe or {})])
    return _html_detalhe_conteudo(item, detalhe)


def _linhas_diff_para_planilha(texto: str, tipo: str) -> dict[str, str]:
    """Normaliza evidência em colunas da planilha Antes/Depois."""
    from backend.app.services.comparador_arquivos import _formatar_valor_planilha

    item = _parse_evidencia_item(str(texto))
    local = str(item.get("local") or ("Inclusão" if tipo == "entrou" else "Alteração"))
    antes = _formatar_valor_planilha(item.get("antes")) if item.get("antes") not in (None, "") else ("—" if tipo == "entrou" else "em branco")
    if item.get("depois") not in (None, ""):
        depois = _formatar_valor_planilha(item.get("depois"))
    else:
        depois = str(texto)
    if depois == "em branco" and tipo == "saiu":
        depois = "—"
    mudanca = str(item.get("mudanca") or "").strip()
    if not mudanca:
        try:
            from backend.app.services.comparador_arquivos import _descrever_mudanca

            if tipo == "entrou":
                mudanca = "acrescentou aba"
            elif tipo == "saiu":
                mudanca = "removeu aba"
            elif "renomeou aba" in str(item.get("mudanca") or ""):
                mudanca = "renomeou aba"
            else:
                mudanca = _descrever_mudanca(
                    "" if antes in {"—", "em branco"} else antes,
                    "" if depois in {"—", "em branco"} else depois,
                )
        except Exception:
            mudanca = "texto alterado"
    return {
        "tipo": {"entrou": "Entrou", "saiu": "Saiu", "mudou": "Mudou"}.get(tipo, tipo),
        "onde": local,
        "mudanca": mudanca,
        "antes": antes if tipo != "entrou" else "—",
        "depois": depois if tipo != "saiu" else "—",
    }


def gerar_planilha_antes_depois(
    alterados: list[dict],
    detalhes_por_url: dict,
) -> tuple[bytes, str] | None:
    """Planilha anexa com TODAS as linhas Antes/Depois + O que fazer (itens que precisam agir)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception as exc:
        logger.warning("openpyxl indisponível para anexo Antes/Depois: %s", exc)
        return None

    precisa: list[tuple[dict, dict]] = []
    for item in alterados:
        det = detalhes_por_url.get(item["url"]) or {}
        if not _detalhe_so_tecnico(det, item.get("evidencia") or ""):
            precisa.append((item, det))
    if not precisa:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Antes_Depois"
    cab = [
        "Leiaute",
        "Arquivo",
        "Tipo de comparação",
        "Tipo",
        "Onde",
        "O que mudou",
        "Antes",
        "Depois",
        "O que fazer",
        "Link Bacen",
    ]
    ws.append(cab)
    header_fill = PatternFill("solid", fgColor="2E3192")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for item, det in precisa:
        o_que_fazer = _texto_o_que_fazer(item, det)
        _cod_cmp, rotulo_cmp = _tipo_comparacao_rotulo(det)
        codigo = str(det.get("leiaute_codigo") or "").strip()
        nome = str(det.get("nome_arquivo") or _filename_from_url(item["url"]))
        url = item["url"]
        incluidos = [
            i
            for i in (det.get("itens_incluidos") or [])
            if "novo arquivo na página" not in str(i).lower()
        ]
        removidos = det.get("itens_removidos") or []
        _, alterados_txt = _separar_itens_tecnicos_e_conteudo(
            det.get("itens_alterados") or []
        )
        alterados_txt = [
            c
            for c in alterados_txt
            if "novo arquivo observado" not in str(c).lower()
        ]
        # Abas: Entrou → Saiu → Mudou (renomeações antes das células).
        blocos = (
            [("entrou", x) for x in incluidos]
            + [("saiu", x) for x in removidos]
            + [("mudou", x) for x in alterados_txt]
        )
        if not blocos:
            ws.append(
                [
                    codigo,
                    nome,
                    rotulo_cmp,
                    "Arquivo novo",
                    "—",
                    "arquivo novo sem diff",
                    "—",
                    "—",
                    o_que_fazer,
                    url,
                ]
            )
            continue
        for tipo, texto in blocos:
            linha = _linhas_diff_para_planilha(str(texto), tipo)
            ws.append(
                [
                    codigo,
                    nome,
                    rotulo_cmp,
                    linha["tipo"],
                    linha["onde"],
                    linha["mudanca"],
                    linha["antes"],
                    linha["depois"],
                    o_que_fazer,
                    url,
                ]
            )

    widths = {
        "A": 12,
        "B": 42,
        "C": 36,
        "D": 14,
        "E": 36,
        "F": 28,
        "G": 40,
        "H": 40,
        "I": 48,
        "J": 28,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.auto_filter.ref = f"A1:J{ws.max_row}"
    ws.freeze_panes = "A2"

    from io import BytesIO

    buf = BytesIO()
    wb.save(buf)
    nome = f"{ANEXO_ANTES_DEPOIS_PREFIXO}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return buf.getvalue(), nome


def montar_corpo_email_alteracoes(
    alterados: list[dict],
    detalhes_por_url: dict,
    categoria_por_url: dict | None = None,
) -> str:
    """Dois blocos limpos: Precisa agir (conteúdo) e Não precisa agir (técnico)."""
    del categoria_por_url
    precisa_agir: list[tuple[dict, dict]] = []
    nao_precisa: list[tuple[dict, dict]] = []

    for item in alterados:
        url = item["url"]
        detalhe = detalhes_por_url.get(url) or {}
        evidencia = item.get("evidencia") or ""
        par = (item, detalhe)
        if _detalhe_so_tecnico(detalhe, evidencia):
            nao_precisa.append(par)
        else:
            precisa_agir.append(par)

    n = len(alterados)
    n_acao = len(precisa_agir)
    n_tech = len(nao_precisa)
    n_pareado = 0
    n_mesmo = 0
    n_sem_ant = 0
    for _item, detalhe in precisa_agir:
        tipo_cmp, _ = _tipo_comparacao_rotulo(detalhe)
        if tipo_cmp == "versao_pareada":
            n_pareado += 1
        elif tipo_cmp == "sem_anterior":
            n_sem_ant += 1
        else:
            n_mesmo += 1

    bloco_acao = ""
    if precisa_agir:
        detalhes = "".join(
            _html_detalhe_conteudo(item, detalhe or None)
            for item, detalhe in precisa_agir
        )
        bloco_acao = f"""
      <h2 class="h-acao">Arquivos para revisar ({n_acao})</h2>
      <p class="desc">
        Cada arquivo abaixo está em 3 passos: situação, diferenças e ação.
        A lista completa também está na planilha em anexo
        <strong>{ANEXO_ANTES_DEPOIS_PREFIXO}_….xlsx</strong>.
      </p>
      {detalhes}
    """

    bloco_tech = ""
    if nao_precisa:
        bloco_tech = f"""
      <h2 class="h-tech">Só aviso (sem ação) ({n_tech})</h2>
      <p class="desc">
        O Bacen republicou o arquivo no site, sem mudança de texto, célula ou tabela.
        Em geral não exige ajuste de rotina.
      </p>
      {_html_lista_tecnica(nao_precisa)}
    """

    partes_lead: list[str] = []
    if n_pareado:
        partes_lead.append(f"{n_pareado} arquivo(s) novo(s) na página do Bacen")
    if n_mesmo:
        partes_lead.append(f"{n_mesmo} arquivo(s) atualizado(s)")
    if n_sem_ant:
        partes_lead.append(f"{n_sem_ant} arquivo(s) novo(s) sem versão anterior")
    if n_tech:
        partes_lead.append(f"{n_tech} só aviso (sem ação)")
    if not partes_lead:
        partes_lead.append(f"{n} arquivo(s)")

    titulo_lead = " · ".join(partes_lead)
    lead_acao = (
        "revise o(s) passo(s) abaixo"
        if n_acao
        else "nenhum item exige ação de conteúdo"
    )

    return f"""
      <p class="lead">
        <strong>{titulo_lead}</strong>
        <span class="muted"> — {lead_acao}</span>
      </p>
      {bloco_acao}
      {bloco_tech}
    """


def gerar_html_email(conteudo_html: str, data_ref: str, logo_cid: str) -> str:
    return f"""
<!DOCTYPE html>
<html lang="pt-br">
<head>
<meta charset="UTF-8">
<style>
  body {{ font-family: Arial, Helvetica, sans-serif; margin: 0; padding: 24px; color: #222; background: #fff; font-size: 15px; line-height: 1.45; }}
  .wrap {{ max-width: 880px; margin: 0 auto; }}
  .lead {{ margin: 16px 0 20px; font-size: 15px; }}
  .h-acao, .h-tech {{
    font-size: 16px; font-weight: bold; margin: 28px 0 6px; padding: 0 0 6px;
    border-bottom: 2px solid {BLUE_BRAND}; color: {BLUE_BRAND};
  }}
  .h-tech {{ border-bottom-color: #999; color: #555; }}
  .desc {{ margin: 0 0 14px; color: #555; font-size: 14px; }}
  .acao {{
    margin: 0 0 10px; padding: 10px 12px; background: #fff8e6;
    border-left: 4px solid #d4a017; color: #333; font-size: 14px;
  }}
  .passo {{
    margin: 0 0 14px; padding: 0;
  }}
  .passo-titulo {{
    margin: 0 0 6px; font-size: 13px; font-weight: bold;
    color: {BLUE_BRAND}; letter-spacing: 0.01em;
  }}
  .passo-next {{
    margin: 8px 0 0; font-size: 12px; color: #777; font-style: italic;
  }}
  .simples {{
    margin: 10px 0 0; padding: 10px 12px; background: #f3faf5;
    border-left: 4px solid #2f7d4a; color: #222; font-size: 14px;
  }}
  .simples-title {{
    margin: 0 0 6px; font-weight: bold; color: #2f7d4a; font-size: 14px;
  }}
  .simples-list {{ margin: 0 0 0 18px; padding: 0; }}
  .simples-list li {{ margin: 0 0 4px; }}
  .file-block {{ margin: 0 0 20px; padding: 0 0 16px; border-bottom: 1px solid #eee; }}
  .file-block:last-child {{ border-bottom: none; }}
  .file-title {{ margin: 0 0 10px; font-size: 15px; font-weight: bold; color: #222; }}
  .sec-label {{ margin: 10px 0 4px; font-size: 13px; font-weight: bold; color: #222; }}
  .diff-table {{ width: 100%; border-collapse: collapse; font-size: 13px; margin: 0 0 6px; }}
  .diff-table th {{
    text-align: left; background: #f3f3f3; color: #555; font-size: 11px;
    text-transform: uppercase; letter-spacing: 0.02em; padding: 6px 8px; border: 1px solid #e0e0e0;
  }}
  .diff-table td {{ vertical-align: top; padding: 6px 8px; border: 1px solid #e0e0e0; word-break: break-word; }}
  .diff-table .col-local {{ width: 30%; font-weight: bold; color: #333; }}
  .compact-list, .tech-list {{ margin: 0 0 8px 18px; padding: 0; font-size: 14px; }}
  .compact-list li, .tech-list li {{ margin: 0 0 4px; }}
  .more {{ color: {BLUE_BRAND}; font-size: 12px; font-weight: bold; margin: 4px 0 0; }}
  .muted {{ color: #777; font-weight: normal; font-size: 13px; }}
  a {{ color: {BLUE_BRAND}; text-decoration: none; }}
  a:hover {{ text-decoration: underline; }}
  .rodape {{ font-size: 12px; color: #777; margin-top: 32px; text-align: center; border-top: 1px solid #eee; padding-top: 16px; }}
</style>
</head>
<body>
  <div class="wrap">
    <div style="text-align:center; margin-bottom: 8px;">
      <img src="cid:{logo_cid}" alt="FINAUD TEC" style="max-width:180px; height:auto;">
    </div>
    <p style="margin: 12px 0 0; font-size: 15px; text-align: center;">
      Atualizações nos leiautes do Bacen em <strong>{data_ref}</strong>.
    </p>
    {conteudo_html}
    <div class="rodape">
      E-mail automático — FINAUD TEC SOLUÇÕES EM TECNOLOGIA
    </div>
  </div>
</body>
</html>
""".strip()

def gerar_html_sem_novidade(data_ref: str, logo_cid: str) -> str:
    return f"""
<!DOCTYPE html>
<html lang="pt-br">
<head>
<meta charset="UTF-8">
<style>
  body {{ font-family: Arial, sans-serif; margin: 20px; color:#111; }}
  .wrap {{ width: 100%; margin: 0 auto; }}
  .rodape {{ font-size: 12px; color: #555; margin-top: 40px; text-align: center; }}
  .msg-sem-novidade {{ text-align: center; color: {BLUE_BRAND}; font-size: 18px; margin-top: 24px; line-height:1.55; }}
</style>
</head>
<body>
  <div class="wrap">
    <div style="text-align:center; margin-bottom: 12px;">
      <img src="cid:{logo_cid}" alt="FINAUD TEC" style="max-width:220px; height:auto;">
    </div>
    <p class="msg-sem-novidade">
      <strong>Nenhum documento novo ou alterado foi identificado na página de leiautes em {data_ref}</strong>.
    </p>
    <div class="rodape">
      Este e-mail foi gerado automaticamente pelo sistema de monitoramento <b>FINAUD TEC SOLUÇÕES EM TECNOLOGIA</b>.
    </div>
  </div>
</body>
</html>
""".strip()


# ====== NOMES ======
def _filename_from_url(url):
    base = unquote(urlparse(url).path).split("/")[-1].strip() or "arquivo"
    return base

def nome_doc_por_url(url):
    if "DDR" in url: return "DDR (2011)"
    if "DRM" in url: return "DRM"
    if "2061" in url: return "DOC 2061"
    if "2062" in url: return "DOC 2062"
    if "2160" in url: return "DRL 2160"
    if "4111" in url: return "4111 - SCD"
    return "Desconhecido"


# ====== DOWNLOAD P/ ANEXO ======
def baixar_para_anexo(session, url, max_single=MAX_SINGLE_ATTACH_SIZE):
    try:
        hi = head_info(session, url)
        cl = hi.get("content_length")
        if cl and cl.isdigit() and int(cl) > max_single:
            return None, None, None, f"pula: Content-Length {cl} > limite"
    except Exception:
        pass

    r = session.get(url, stream=True, allow_redirects=True, timeout=TIMEOUT)
    if r.status_code != 200:
        return None, None, None, f"status {r.status_code}"

    data, total = bytearray(), 0
    for chunk in r.iter_content(64 * 1024):
        if not chunk: break
        data.extend(chunk); total += len(chunk)
        if total > max_single:
            return None, None, None, f"pula: excedeu {max_single} bytes"
    content = bytes(data)

    filename = _filename_from_url(r.url or url)
    ctype, _ = mimetypes.guess_type(filename)
    if not ctype: ctype = "application/octet-stream"
    maintype, subtype = ctype.split("/", 1)
    return content, maintype, subtype, None


# ====== CONFIG DE E-MAIL ======
def load_email_config(path: Path):
    cfg = json.loads(path.read_text(encoding="utf-8"))
    to: list[str] = []
    try:
        from persistencia.usuarios_db import listar_emails_alerta

        to = listar_emails_alerta()
        logger.info(
            "Destinatários pelo cadastro (ativos + alerta): %s",
            ", ".join(to) if to else "(nenhum)",
        )
    except Exception as e:
        logger.warning("Falha ao ler destinatários dos usuários: %s", e)

    test_to = os.environ.get("LEIAUTES_EMAIL_TEST_TO", "").strip()
    if test_to:
        to = [x.strip() for x in re.split(r"[,;]", test_to) if x.strip()]
        logger.info(
            "LEIAUTES_EMAIL_TEST_TO ativo: redirecionando envio para %s",
            ", ".join(to),
        )
    smtp = cfg.get("smtp", {})
    return {
        "from": cfg.get("from") or cfg.get("user"),
        "to": to,
        "user": cfg.get("user") or smtp.get("user"),
        "password": cfg.get("password") or cfg.get("senha_app") or cfg.get("pass"),
        "host": smtp.get("host", "smtp.gmail.com"),
        "port": int(smtp.get("port", 465)),
        "ssl": bool(smtp.get("ssl", True)),
        "tls": bool(smtp.get("tls", False)),
    }


# ===== DEFINIÇÃO DA MAIN =====
def main():
    global ONLY_ATUAL
    logger.info("Iniciando monitoração...")
    execucao_id_env = os.environ.get("LEIAUTES_EXECUCAO_ID", "").strip()
    execucao_id = int(execucao_id_env) if execucao_id_env.isdigit() else None
    disable_email = os.environ.get("LEIAUTES_DISABLE_EMAIL", "").strip().lower() in (
        "1",
        "true",
        "yes",
    )
    quiet_cfg, ONLY_ATUAL = _carregar_flags_monitor()
    logger.info(
        "Flags monitor: quiet_baseline=%s only_atual=%s disable_email=%s",
        quiet_cfg,
        ONLY_ATUAL,
        disable_email,
    )

    anexos_detectados = []
    categoria_por_url = {}
    links_detectados_por_data = []

    for url in urls:
        try:
            datas, anexos, categorias = extrair_datas_categorias_e_anexos(url)
            if hoje in datas:
                links_detectados_por_data.append(url)
            for link in anexos:
                anexos_detectados.append(link)
                categoria_por_url[link] = categorias.get(link, "Sem categoria")
        except Exception as e:
            logger.warning(f"Erro ao processar URL {url}: {e}")
            continue

    # Se o scrape Angular falhar em alguma página, ainda recheca o que já conhecemos
    # (metadados). Arquivo novo na página continua dependendo do scrape.
    vistos = set(anexos_detectados)
    for url_conhecida, _meta in _load_manifest().items():
        if url_conhecida in vistos:
            continue
        anexos_detectados.append(url_conhecida)
        vistos.add(url_conhecida)
        categoria_por_url.setdefault(url_conhecida, _categoria_da_pagina(url_conhecida))
    logger.info(
        "Anexos para checagem: %s (scrape + fallback manifesto).",
        len(anexos_detectados),
    )

    alterados, manifest, flags_anexos = verificar_anexos(
        anexos_detectados,
        categoria_por_url=categoria_por_url,
        execucao_id=execucao_id,
    )
    anexos_nomes = [_filename_from_url(a["url"]) for a in alterados]
    quiet_baseline = bool(flags_anexos.get("quiet_baseline"))
    first_run = bool(flags_anexos.get("first_run"))

    emails_enviados = 0
    destinatarios = []
    enviar_sem_alteracao, anexar_alterados = _carregar_flags_email()
    logger.info(
        "Flags e-mail: enviar_sem_alteracao=%s anexar_alterados=%s",
        enviar_sem_alteracao,
        anexar_alterados,
    )

    if disable_email:
        logger.info("Envio de e-mail desativado por LEIAUTES_DISABLE_EMAIL=1.")
    elif first_run and quiet_baseline and not alterados:
        logger.info(
            "Primeira baseline quieta (monitor.quiet_baseline=true): e-mail omitido."
        )
    elif alterados or enviar_sem_alteracao:
        email_cfg = load_email_config(CONFIG_PATH)
        destinatarios = email_cfg.get("to", [])
        if not destinatarios:
            logger.warning(
                "Nenhum destinatário: cadastre usuários ativos com "
                "'Receber e-mail de alertas'."
            )
        else:
            msg = MIMEMultipart()
            msg["Subject"] = ASSUNTO
            msg["From"] = email_cfg["from"]
            msg["To"] = ", ".join(destinatarios)

            logo_cid = make_msgid(domain="finaud.com.br")[1:-1]
            detalhes_por_url: dict = {}

            if alterados:
                detalhes_por_url = _carregar_detalhes_alteracoes(execucao_id)
                corpo = montar_corpo_email_alteracoes(
                    alterados,
                    detalhes_por_url,
                    categoria_por_url,
                )
                corpo_html = gerar_html_email(corpo, hoje, logo_cid)
            else:
                corpo_html = gerar_html_sem_novidade(hoje, logo_cid)

            msg.attach(MIMEText(corpo_html, "html", "utf-8"))

            with open(LOGO_PATH, "rb") as f:
                img = MIMEImage(f.read())
                img.add_header("Content-ID", f"<{logo_cid}>")
                img.add_header("Content-Disposition", "inline", filename="logo.jpg")
                msg.attach(img)

            session = _session()
            total_size = 0
            anexados = 0
            # 1) Planilha Antes/Depois completa — é o anexo principal do alerta.
            if alterados:
                planilha = gerar_planilha_antes_depois(alterados, detalhes_por_url)
                if planilha:
                    content_xlsx, nome_xlsx = planilha
                    part = MIMEBase(
                        "application",
                        "vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                    part.set_payload(content_xlsx)
                    encoders.encode_base64(part)
                    part.add_header(
                        "Content-Disposition", "attachment", filename=nome_xlsx
                    )
                    msg.attach(part)
                    total_size += len(content_xlsx)
                    anexados += 1
                    logger.info(
                        "Anexo principal Antes/Depois: %s (%s bytes)",
                        nome_xlsx,
                        len(content_xlsx),
                    )
                else:
                    logger.warning(
                        "Planilha Antes/Depois não gerada (openpyxl ou sem itens)."
                    )
            # 2) Arquivos Bacen alterados (se flag da tela permitir).
            bacen_anexados = 0
            if alterados and anexar_alterados:
                for item in alterados[:MAX_ATTACHMENTS]:
                    url = item["url"]
                    content, maintype, subtype, motivo = baixar_para_anexo(
                        session, url
                    )
                    if content and maintype and subtype:
                        if total_size + len(content) > MAX_TOTAL_ATTACH_SIZE:
                            logger.warning(
                                "Limite total de anexos atingido; pulando restantes."
                            )
                            break
                        filename = _filename_from_url(url)
                        part = MIMEBase(maintype, subtype)
                        part.set_payload(content)
                        encoders.encode_base64(part)
                        part.add_header(
                            "Content-Disposition",
                            "attachment",
                            filename=filename,
                        )
                        msg.attach(part)
                        total_size += len(content)
                        anexados += 1
                        bacen_anexados += 1
                    elif motivo:
                        logger.warning(
                            f"Não foi possível anexar {url} | Motivo: {motivo}"
                        )
            elif alterados and not anexar_alterados:
                logger.info(
                    "Anexos Bacen omitidos (email.anexar_alterados=false); "
                    "planilha Antes/Depois mantida."
                )

            try:
                smtp_class = smtplib.SMTP_SSL if email_cfg["ssl"] else smtplib.SMTP
                with smtp_class(email_cfg["host"], email_cfg["port"]) as server:
                    if email_cfg["tls"]:
                        server.starttls()
                    if email_cfg["user"] and email_cfg["password"]:
                        server.login(email_cfg["user"], email_cfg["password"])
                    server.sendmail(email_cfg["from"], destinatarios, msg.as_string())
                logger.info(f"E-mail enviado para: {', '.join(destinatarios)}")
                emails_enviados = 1
            except Exception as e:
                logger.error(f"Erro ao enviar e-mail: {e}")
    else:
        logger.info(
            "E-mail omitido: sem alterações e email.enviar_sem_alteracao=false."
        )

    return {
        "links_detectados_por_data": links_detectados_por_data,
        "alterados": alterados,
        "emails_enviados": emails_enviados,
        "destinatarios": destinatarios,
        "anexos_nomes": anexos_nomes,
    }


# ===== Helpers =====
def _plural(n: int, sing: str, plur: str | None = None) -> str:
    if n == 1:
        return f"{n} {sing}"
    return f"{n} {plur or (sing + 's')}"

def _fmt_duracao(delta: timedelta) -> str:
    s = int(delta.total_seconds())
    h, rem = divmod(s, 3600)
    m, s = divmod(rem, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


# ===== MAIN RUN =====
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Monitor de leiautes Bacen")
    parser.add_argument(
        "--checar-agenda",
        action="store_true",
        help="Só executa se horário/dia estiver na agenda do banco (cron frequente).",
    )
    parser.add_argument(
        "--ignorar-agenda",
        action="store_true",
        help="Ignora checagem de agenda (execução manual / testes).",
    )
    args = parser.parse_args()

    if args.checar_agenda and not args.ignorar_agenda:
        try:
            from persistencia.agenda_db import deve_executar_agora

            pode, motivo = deve_executar_agora()
            if not pode:
                print(f"[agenda] Não executa: {motivo}")
                raise SystemExit(0)
            print(f"[agenda] Executa: {motivo}")
        except SystemExit:
            raise
        except Exception as exc:
            print(f"[agenda] Falha ao ler agenda ({exc}); não executa por segurança.")
            raise SystemExit(0)

    # Cron/CLI: cria execução no banco se a API não passou LEIAUTES_EXECUCAO_ID.
    # Sem isso, versões/alterações (antes/depois) não são gravadas e o e-mail fica só com metadados.
    execucao_propria = False
    execucao_id_cli = None
    env_exec = os.environ.get("LEIAUTES_EXECUCAO_ID", "").strip()
    if not env_exec.isdigit() and iniciar_execucao:
        try:
            execucao_id_cli = iniciar_execucao()
            os.environ["LEIAUTES_EXECUCAO_ID"] = str(execucao_id_cli)
            execucao_propria = True
            logger.info(
                "Execução %s criada pelo motor (cron/CLI) para evidências antes/depois.",
                execucao_id_cli,
            )
        except Exception as exc:
            logger.warning("Não foi possível criar execução no banco: %s", exc)

    try:
        inicio_exec = datetime.now()
        result = main()
        fim_exec = datetime.now()
        duracao = _fmt_duracao(fim_exec - inicio_exec)

        links_detectados_por_data = result.get("links_detectados_por_data", [])
        alterados = result.get("alterados", [])
        urls_alterados = [a["url"] for a in alterados]
        links_com_alteracao = not any(
            any(link.startswith(url_base) for url_base in links_detectados_por_data)
            for link in urls_alterados
        )
        emails_enviados = int(result.get("emails_enviados", 0))
        destinatarios = result.get("destinatarios", [])
        anexos_nomes = result.get("anexos_nomes", [])
        destinatarios_str = ", ".join(destinatarios) if destinatarios else "—"

        paginas_verificadas = len(urls)
        leiautes_novos = len(links_detectados_por_data)
        pdfs_gerados = len(alterados)

        logger.info(f"Total de links únicos detectados: {leiautes_novos}")
        logger.info(f"Documentos alterados: {pdfs_gerados}")
        logger.info(f"PDFs gerados: {len(anexos_nomes)}")
        if anexos_nomes:
            logger.info("Documentos gerados:")
            for nome in anexos_nomes:
                logger.info(f" - {nome}")
        else:
            logger.info("Nenhum PDF foi gerado.")

        logger.info(f"E-mails enviados: {emails_enviados}")
        logger.info(f"Destinatários: {destinatarios_str}")

        txt_leiautes = _plural(leiautes_novos, "leiaute novo", "leiautes novos")
        txt_pdfs     = _plural(pdfs_gerados,    "PDF gerado",   "PDFs gerados")
        
        aviso_tecnico = ""
        
        if leiautes_novos > 0:
            header = f"🟢 OK | {txt_leiautes}, {txt_pdfs} | em {duracao}"
            if leiautes_novos > 0:
                exemplo_nome = nome_doc_por_url(links_detectados_por_data[0])
                header += f" | ex: {exemplo_nome}"
        elif alterados:
            header = f"🟡 AVISO | Link(s) alterado(s), sem data nova | em {duracao}"
            aviso_tecnico = "🛈 AVISO TÉCNICO: link(s) foram alterados no Bacen, mesmo sem data nova"
        else:
            header = f"🟢 OK | Nenhuma alteração detectada | em {duracao}"

        # Gera nomes legíveis dos leiautes verificados
        codigo_para_sigla = {
            "2061": "DLO",
            "2062": "DLI",
            "DRM": "DRM",
            "2011": "DDR",
            "2160": "DRL",
            "2060": "DRM",
            "4111": "SCD",
        
        }
        
        paginas_formatadas = []
        for url in urls:
            encontrado = False
            for codigo, sigla in codigo_para_sigla.items():
                if codigo in url.upper():
                    if sigla == "SCD":
                        paginas_formatadas.append(f"{sigla} - 4111")
                    if codigo == "DRM":
                        paginas_formatadas.append("DRM - 2060")
                    else:
                        paginas_formatadas.append(f"{sigla} - {codigo}")
                    encontrado = True
                    break
            if not encontrado:
                trecho = url.rsplit("/", 1)[-1].upper().replace("LEIAUTEDOCUMENTO", "").replace("LEIAUTEDOC", "")
                if "SCRD" in trecho:  # <- correção
                    paginas_formatadas.append("SCD - 4111")
                else:
                    paginas_formatadas.append(f"{trecho} - (desconhecido)")        
        
        if len(paginas_formatadas) > 1:
            leiautes_str = ", ".join(paginas_formatadas[:-1]) + " e " + paginas_formatadas[-1]
        else:
            leiautes_str = paginas_formatadas[0]
        
        resumo = {
            "📄 Leiautes verificados": leiautes_str,
            "📊 Leiautes novos": leiautes_novos,
            "📄 PDFs gerados": pdfs_gerados,
            "📧 E-mails enviados": emails_enviados,
            "✉️ Destinatários": destinatarios_str,
            "📄 Arquivos com mudanças detectadas": "\n- " + "\n- ".join(anexos_nomes) if anexos_nomes else "Nenhum",
        }

        _write_status_tail("leiautes", header, resumo, [], aviso_tecnico)

        if execucao_propria and execucao_id_cli and finalizar_execucao:
            contadores = (
                contar_resultados_execucao(execucao_id_cli)
                if contar_resultados_execucao
                else {"qtd_leiautes": 0, "qtd_arquivos": 0, "qtd_alteracoes": len(alterados)}
            )
            finalizar_execucao(
                execucao_id_cli,
                status="sucesso",
                qtd_leiautes=contadores.get("qtd_leiautes", 0),
                qtd_arquivos=contadores.get("qtd_arquivos", 0),
                qtd_alteracoes=contadores.get("qtd_alteracoes", len(alterados)),
                emails_enviados=emails_enviados,
            )

    except Exception as e:
        if execucao_propria and execucao_id_cli and finalizar_execucao:
            try:
                finalizar_execucao(
                    execucao_id_cli,
                    status="erro",
                    erro=str(e).splitlines()[-1][:2000],
                )
            except Exception:
                pass
        try:
            resumo_err = {"Motivo": str(e).splitlines()[-1]}
        except Exception:
            resumo_err = {"Motivo": str(e)}
        try:
            extra = "Veja o log para o traceback completo."
            _write_status_tail("leiautes", "🔴 ERRO | Falha na execução", resumo_err, [], extra)
        except Exception as log_error:
            print("Falha ao escrever no status_tail:", log_error)
        raise
    finally:
        print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | === FIM leiautes ===")
